"""Write a plain .xlsx table to a new file, using the shared instruction entry."""

from __future__ import annotations

from copy import copy
from datetime import datetime, time, timedelta
from io import BytesIO
from math import isclose, isfinite
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension
from openpyxl.utils.exceptions import IllegalCharacterError
from pydantic import Field

from .core import Instruction, InstructionError, InstructionModel
from .excel import CellValue


INSTRUCTION_ID = "excel.write"


class WriteExcelInput(InstructionModel):
    file_path: str = Field(min_length=1, description="新建 .xlsx 文件的保存路径，所在目录须已存在")
    columns: list[str] = Field(min_length=1, max_length=16384, description="按保存顺序填写列名")
    rows: list[dict[str, CellValue]] = Field(max_length=1048575, description="每行按列名保存数据；键须与列名一致")
    sheet_name: str = Field(default="数据", min_length=1, max_length=31, description="工作表名称")
    template_file: str | None = Field(default=None, min_length=1, description="可选原 .xlsx；保留原工作簿，仅在指定表右侧追加列，原列和全部有效行必须原样传入")


class WriteExcelOutput(InstructionModel):
    file_path: str = Field(description="生成文件的绝对路径")
    sheet_name: str = Field(description="生成的工作表名称")
    row_count: int = Field(ge=0, description="写入的数据行数，不含表头；含传入的全空行")


def _failure(code: str, message: str) -> InstructionError:
    return InstructionError(code, INSTRUCTION_ID, "execute", message)


def _check_table(inputs: WriteExcelInput) -> None:
    if (
        not inputs.sheet_name.strip()
        or any(character in inputs.sheet_name for character in "\\/*?:[]")
        or inputs.sheet_name.startswith("'")
        or inputs.sheet_name.endswith("'")
    ):
        raise _failure("EXCEL_SHEET_INVALID", "工作表名称无效，请使用普通文字名称。")
    if any(not column or column != column.strip() for column in inputs.columns):
        raise _failure("EXCEL_HEADER_INVALID", "列名不能为空，也不能在首尾带空格。")
    expected = set(inputs.columns)
    if len(expected) != len(inputs.columns):
        raise _failure("EXCEL_HEADER_INVALID", "列名不能重复。")
    for row_number, row in enumerate(inputs.rows, start=2):
        if set(row) != expected:
            raise _failure("EXCEL_ROW_INVALID", f"第 {row_number} 行的字段与列名不一致，请补齐缺列或移除多余字段。")


def _set_cell(worksheet, row: int, column: int, value: CellValue) -> None:
    location = f"第 {row} 行、第 {column} 列"
    if isinstance(value, str) and len(value) > 32767:
        raise _failure("EXCEL_VALUE_INVALID", f"{location}的文字超过 Excel 单元格长度限制。")
    if isinstance(value, float) and not isfinite(value):
        raise _failure("EXCEL_VALUE_INVALID", f"{location}不是有限数字。")
    if type(value) is int and abs(value) > 999999999999999:
        raise _failure("EXCEL_VALUE_INVALID", f"{location}的整数超过 15 位，请以文字传入以保留完整内容。")
    if isinstance(value, (datetime, time)) and value.tzinfo is not None:
        raise _failure("EXCEL_VALUE_INVALID", f"{location}的日期或时间带有时区，请先转换为所需的本地时间。")
    if isinstance(value, (datetime, time, timedelta)):
        microseconds = value.microseconds if isinstance(value, timedelta) else value.microsecond
        if microseconds % 1000:
            raise _failure("EXCEL_VALUE_INVALID", f"{location}的时间精度小于毫秒，请先明确需要保留的精度。")
    try:
        cell = worksheet.cell(row=row, column=column, value=value)
        if isinstance(value, str):
            # Preserve literal strings, including leading '=' and Excel error names.
            cell.data_type = "s"
    except (IllegalCharacterError, TypeError, ValueError) as exc:
        raise _failure("EXCEL_VALUE_INVALID", f"{location}包含无法写入 Excel 的内容。") from exc


def _load_template(path: str):
    source = Path(path)
    if source.suffix.lower() != ".xlsx":
        raise _failure("EXCEL_FORMAT_UNSUPPORTED", "原表必须是 .xlsx 文件。")
    try:
        with source.open("rb") as stream:
            return load_workbook(stream, data_only=False, keep_links=True)
    except (OSError, BadZipFile, ValueError, KeyError) as exc:
        raise _failure("EXCEL_TEMPLATE_INVALID", "无法读取原工作簿，请检查文件和格式。") from exc


def _template_records(workbook, inputs: WriteExcelInput) -> tuple[int, list[int]]:
    if inputs.sheet_name not in workbook.sheetnames:
        raise _failure("EXCEL_SHEET_MISSING", "原工作簿中没有指定的工作表。")
    sheet = workbook[inputs.sheet_name]
    headers = [cell.value for cell in sheet[1]]
    while headers and headers[-1] in (None, ""):
        headers.pop()
    if not headers or any(not isinstance(value, str) or not value.strip() for value in headers):
        raise _failure("EXCEL_TEMPLATE_INVALID", "原表第一行必须是连续的非空文字列名。")
    columns = [value.strip() for value in headers]
    if (len(set(columns)) != len(columns) or inputs.columns[:len(columns)] != columns
            or len(inputs.columns) <= len(columns)):
        raise _failure("EXCEL_TEMPLATE_MISMATCH", "保留原表时只能在右侧追加新列，原列名及顺序必须一致。")
    if any(merged.max_col > len(columns) and merged.min_col <= len(inputs.columns)
           for merged in sheet.merged_cells.ranges):
        raise _failure("EXCEL_TEMPLATE_INVALID", "新增列与原表合并单元格重叠，请先调整原表。")
    row_numbers = []
    for cells in sheet.iter_rows(min_row=2):
        if any(cell.data_type in {"f", "e"} for cell in cells):
            raise _failure("EXCEL_CELL_UNSUPPORTED", "需要追加列的工作表含公式或错误值，请先转为有效数据。")
        values = [cell.value for cell in cells]
        if all(value is None or value == "" for value in values):
            continue
        if any(value is not None and value != "" for value in values[len(columns):]):
            raise _failure("EXCEL_TEMPLATE_INVALID", "原表存在没有列名的数据，请补齐表头。")
        index = len(row_numbers)
        if index >= len(inputs.rows) or any(
            type(inputs.rows[index][column]) is not type(value)
            or inputs.rows[index][column] != value
            for column, value in zip(columns, values)
        ):
            raise _failure("EXCEL_TEMPLATE_MISMATCH", "原列数据或行顺序发生变化，追加列时必须保留全部有效行。")
        row_numbers.append(cells[0].row)
    if len(row_numbers) != len(inputs.rows):
        raise _failure("EXCEL_TEMPLATE_MISMATCH", "数据行数与原表不一致，不能通过追加列删除或新增数据行。")
    return len(columns), row_numbers


def _append_columns(workbook, inputs: WriteExcelInput) -> None:
    original_count, row_numbers = _template_records(workbook, inputs)
    sheet = workbook[inputs.sheet_name]
    for column, name in enumerate(inputs.columns[original_count:], start=original_count + 1):
        header = sheet.cell(1, column)
        header._style = copy(sheet.cell(1, original_count)._style)
        _set_cell(sheet, 1, column, name)
        sheet.column_dimensions[get_column_letter(column)].width = max(12, min(40, len(name) * 2 + 2))
        for row_number, row in zip(row_numbers, inputs.rows):
            cell = sheet.cell(row_number, column)
            cell._style = copy(sheet.cell(row_number, original_count)._style)
            # Keep neighboring typography but let new values retain their own precision/type.
            cell.number_format = "General"
            _set_cell(sheet, row_number, column, row[name])


def _build_workbook(inputs: WriteExcelInput) -> bytes:
    if inputs.template_file is not None:
        workbook = _load_template(inputs.template_file)
        try:
            _append_columns(workbook, inputs)
            workbook.iso_dates = True
            with BytesIO() as buffer:
                workbook.save(buffer)
                return buffer.getvalue()
        finally:
            workbook.close()
    workbook = Workbook(iso_dates=True)
    try:
        worksheet = workbook.active
        worksheet.title = inputs.sheet_name
        for column, name in enumerate(inputs.columns, start=1):
            _set_cell(worksheet, 1, column, name)
        for row_number, row in enumerate(inputs.rows, start=2):
            for column, name in enumerate(inputs.columns, start=1):
                _set_cell(worksheet, row_number, column, row[name])
        with BytesIO() as buffer:
            workbook.save(buffer)
            return buffer.getvalue()
    except InstructionError:
        raise
    except (TypeError, ValueError) as exc:
        raise _failure("EXCEL_VALUE_INVALID", "工作表名称或数据无法保存为 Excel，请检查输入内容。") from exc
    finally:
        workbook.close()


def write_excel(inputs: WriteExcelInput) -> dict[str, object]:
    path = Path(inputs.file_path).absolute()
    if path.suffix.lower() != ".xlsx":
        raise _failure("EXCEL_FORMAT_UNSUPPORTED", "保存路径必须以 .xlsx 结尾。")
    _check_table(inputs)
    payload = _build_workbook(inputs)
    try:
        destination = path.open("xb")
    except FileExistsError as exc:
        raise _failure("EXCEL_FILE_EXISTS", "目标文件已存在，请换一个保存名称。") from exc
    except (FileNotFoundError, NotADirectoryError) as exc:
        raise _failure("EXCEL_DIRECTORY_MISSING", "保存目录不存在，请先创建目录或更换路径。") from exc
    except OSError as exc:
        raise _failure("EXCEL_WRITE_FAILED", "无法创建文件，请检查保存路径和目录权限。") from exc

    try:
        with destination:
            destination.write(payload)
    except BaseException as exc:
        # Only this invocation's newly created file can reach this cleanup.
        try:
            path.unlink()
        except OSError as cleanup_error:
            if not isinstance(exc, Exception):
                exc.add_note("本次新建的目标文件未能清理，请检查残留文件。")
                raise exc from cleanup_error
            raise _failure("EXCEL_WRITE_INCOMPLETE", "写入未完成，目标位置残留本次新建文件，请检查后再处理。") from cleanup_error
        if not isinstance(exc, Exception):
            raise
        raise _failure("EXCEL_WRITE_FAILED", "写入失败，本次未完成的新文件已清理。") from exc
    return {"file_path": str(path), "sheet_name": inputs.sheet_name, "row_count": len(inputs.rows)}


def verify_excel_write(inputs: WriteExcelInput, result: WriteExcelOutput) -> bool:
    if (
        result.file_path != str(Path(inputs.file_path).absolute())
        or result.sheet_name != inputs.sheet_name
        or result.row_count != len(inputs.rows)
    ):
        return False
    if inputs.template_file is not None:
        return _verify_template_write(inputs, result)
    with Path(result.file_path).open("rb") as source:
        workbook = load_workbook(source, read_only=True, data_only=False, keep_links=False)
        try:
            if workbook.sheetnames != [inputs.sheet_name]:
                return False
            worksheet = workbook[inputs.sheet_name]
            return (
                worksheet.max_row == len(inputs.rows) + 1
                and worksheet.max_column == len(inputs.columns)
                and [cell.value for cell in worksheet[1]] == inputs.columns
            )
        finally:
            workbook.close()


def _same_saved_value(expected: CellValue, actual: CellValue) -> bool:
    if expected == "":
        expected = None
    if type(expected) in (int, float) and type(actual) in (int, float):
        # A decimal such as 7/3 may round during xlsx serialization; bool is never numeric here.
        return isclose(expected, actual, rel_tol=1e-15, abs_tol=0.0)
    return type(expected) is type(actual) and expected == actual


def _verify_template_write(inputs: WriteExcelInput, result: WriteExcelOutput) -> bool:
    original = _load_template(inputs.template_file)
    try:
        count, row_numbers = _template_records(original, inputs)
        with Path(result.file_path).open("rb") as stream:
            written = load_workbook(stream, data_only=False, keep_links=True)
        try:
            if written.sheetnames != original.sheetnames:
                return False
            additions = {(1, index): name for index, name in enumerate(inputs.columns[count:], start=count + 1)}
            additions.update({(row_number, index): row[name]
                for row_number, row in zip(row_numbers, inputs.rows)
                for index, name in enumerate(inputs.columns[count:], start=count + 1)})
            for old_sheet, new_sheet in zip(original.worksheets, written.worksheets):
                if (str(old_sheet.merged_cells) != str(new_sheet.merged_cells)
                        or old_sheet.freeze_panes != new_sheet.freeze_panes
                        or old_sheet.sheet_state != new_sheet.sheet_state):
                    return False
                # WPS may store stale outline summary levels; openpyxl recalculates them.
                # Actual per-row/column grouping remains checked below.
                if any(getattr(old_sheet.sheet_format, name) != getattr(new_sheet.sheet_format, name)
                       for name in ("baseColWidth", "defaultColWidth", "defaultRowHeight",
                                    "customHeight", "zeroHeight", "thickTop", "thickBottom", "outlineLevelRow")):
                    return False
                for key in set(old_sheet.column_dimensions) | set(new_sheet.column_dimensions):
                    if old_sheet.title == inputs.sheet_name and column_index_from_string(key) > count:
                        continue
                    old = old_sheet.column_dimensions.get(key) or ColumnDimension(old_sheet, index=key)
                    new = new_sheet.column_dimensions.get(key) or ColumnDimension(new_sheet, index=key)
                    if any(getattr(old, name) != getattr(new, name)
                           for name in ("width", "hidden", "bestFit", "outlineLevel", "collapsed")):
                        return False
                for key in set(old_sheet.row_dimensions) | set(new_sheet.row_dimensions):
                    old = old_sheet.row_dimensions.get(key) or RowDimension(old_sheet, index=key)
                    new = new_sheet.row_dimensions.get(key) or RowDimension(new_sheet, index=key)
                    if any(getattr(old, name) != getattr(new, name)
                           for name in ("height", "hidden", "outlineLevel", "collapsed")):
                        return False
                for cells in new_sheet.iter_rows():
                    for cell in cells:
                        key = (cell.row, cell.column)
                        if old_sheet.title == inputs.sheet_name and key in additions:
                            value = additions[key]
                            if not _same_saved_value(value, cell.value):
                                return False
                            if cell.data_type in {"f", "e"}:
                                return False
                        else:
                            previous = old_sheet.cell(cell.row, cell.column)
                            if (cell.value != previous.value or cell.data_type != previous.data_type
                                    or cell.number_format != previous.number_format
                                    or copy(cell.font) != copy(previous.font)
                                    or copy(cell.alignment) != copy(previous.alignment)
                                    or copy(cell.fill) != copy(previous.fill)
                                    or copy(cell.border) != copy(previous.border)):
                                return False
                # A shortened sheet must not hide deleted original cells from the comparison above.
                for cells in old_sheet.iter_rows():
                    for previous in cells:
                        if (old_sheet.title == inputs.sheet_name
                                and (previous.row, previous.column) in additions):
                            continue
                        cell = new_sheet.cell(previous.row, previous.column)
                        if cell.value != previous.value or cell.data_type != previous.data_type:
                            return False
            sheet = written[inputs.sheet_name]
            return all(_same_saved_value(value, sheet.cell(row, col).value) for (row, col), value in additions.items()
                       if value not in (None, ""))
        finally:
            written.close()
    finally:
        original.close()


WRITE_EXCEL = Instruction(
    instruction_id=INSTRUCTION_ID,
    name="写入 Excel",
    version="0.1.1",
    description="写成新的 .xlsx；可基于原工作簿只追加新列，保留原数据与其他工作表。目标文件已存在时拒绝覆盖。",
    input_model=WriteExcelInput,
    output_model=WriteExcelOutput,
    handler=write_excel,
    verifier=verify_excel_write,
)
