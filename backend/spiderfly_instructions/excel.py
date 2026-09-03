"""Read a simple .xlsx table without opening Excel or changing the file."""

from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime, time, timedelta
from pathlib import Path
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import Field

from .core import Instruction, InstructionError, InstructionModel, InstructionRegistry


CellValue = str | bool | int | float | datetime | date | time | timedelta | None
INSTRUCTION_ID = "excel.read"


class ReadExcelInput(InstructionModel):
    file_path: str = Field(min_length=1, description="本地 .xlsx 文件路径")
    sheet_name: str | None = Field(default=None, min_length=1, description="工作表名称；省略时读取第一张")
    required_columns: list[str] = Field(default_factory=list, description="必须存在的列名；忽略列名首尾空白")


class ReadExcelOutput(InstructionModel):
    sheet_name: str = Field(description="实际读取的工作表名称")
    columns: list[str] = Field(description="首行列名，已去除首尾空白")
    rows: list[dict[str, CellValue]] = Field(description="按列名保存的各行数据，不含完全空行")
    row_count: int = Field(ge=0, description="返回的数据行数，不含表头")


def _failure(code: str, message: str) -> InstructionError:
    return InstructionError(code, INSTRUCTION_ID, "execute", message)


def _values(cells: tuple, sheet_name: str) -> list[CellValue]:
    values = []
    for cell in cells:
        if cell.data_type in {"f", "e"}:
            raise _failure(
                "EXCEL_CELL_UNSUPPORTED",
                f"工作表“{sheet_name}”的 {cell.coordinate} 含公式或错误值，请先转为有效数据。",
            )
        values.append(cell.value)
    return values


def _read_sheet(worksheet, inputs: ReadExcelInput) -> dict[str, object]:
    # Read actual cells even when an exporting application wrote wrong dimensions.
    worksheet.reset_dimensions()
    source_rows = worksheet.iter_rows()
    try:
        header = _values(next(source_rows, ()), worksheet.title)
        while header and header[-1] in (None, ""):
            header.pop()
        if not header or any(not isinstance(value, str) or not value.strip() for value in header):
            raise _failure("EXCEL_HEADER_INVALID", "第一行必须是非空文字列名，列名之间不能留空。")
        columns = [value.strip() for value in header]
        if len(set(columns)) != len(columns):
            raise _failure("EXCEL_HEADER_INVALID", "第一行存在重复列名，请修改后再读取。")
        required = [column.strip() for column in inputs.required_columns]
        if any(not column for column in required):
            raise _failure("EXCEL_COLUMNS_INVALID", "必需列名不能为空。")
        missing = [column for column in required if column not in columns]
        if missing:
            raise _failure("EXCEL_COLUMNS_MISSING", "缺少必需列：" + "、".join(missing))

        rows = []
        for row_number, cells in enumerate(source_rows, start=2):
            values = _values(cells, worksheet.title)
            if all(value is None or value == "" for value in values):
                continue
            if any(value is not None and value != "" for value in values[len(columns):]):
                raise _failure(
                    "EXCEL_HEADER_INVALID", f"第 {row_number} 行有数据落在无列名的位置，请补齐第一行列名。"
                )
            values = values[:len(columns)]
            values.extend([None] * (len(columns) - len(values)))
            rows.append(dict(zip(columns, values)))
        return {"sheet_name": worksheet.title, "columns": columns, "rows": rows, "row_count": len(rows)}
    finally:
        source_rows.close()


def read_excel(inputs: ReadExcelInput) -> dict[str, object]:
    path = Path(inputs.file_path)
    if path.suffix.lower() != ".xlsx":
        raise _failure("EXCEL_FORMAT_UNSUPPORTED", "当前只支持 .xlsx 文件，请先转换文件格式。")
    try:
        with path.open("rb") as source:
            workbook = load_workbook(source, read_only=True, data_only=False, keep_links=False)
            try:
                worksheets = workbook.worksheets
                if not worksheets:
                    raise _failure("EXCEL_SHEET_MISSING", "文件中没有可读取的工作表。")
                worksheet = worksheets[0]
                if inputs.sheet_name is not None:
                    worksheet = next(
                        (sheet for sheet in worksheets if sheet.title == inputs.sheet_name), None
                    )
                if worksheet is None:
                    raise _failure("EXCEL_SHEET_MISSING", f"找不到工作表“{inputs.sheet_name}”，请检查名称。")
                return _read_sheet(worksheet, inputs)
            finally:
                workbook.close()
    except (FileNotFoundError, IsADirectoryError) as exc:
        raise _failure("EXCEL_FILE_MISSING", "找不到指定的 Excel 文件，请检查文件路径。") from exc
    except PermissionError as exc:
        raise _failure("EXCEL_ACCESS_DENIED", "无法读取文件，请检查文件访问权限。") from exc
    except (BadZipFile, InvalidFileException, ParseError, KeyError, ValueError) as exc:
        raise _failure("EXCEL_FILE_INVALID", "文件不是有效的 .xlsx 工作簿，请检查文件或重新导出。") from exc


def verify_excel(inputs: ReadExcelInput, result: ReadExcelOutput) -> bool:
    return (
        bool(result.columns)
        and len(set(result.columns)) == len(result.columns)
        and result.row_count == len(result.rows)
        and (inputs.sheet_name is None or inputs.sheet_name == result.sheet_name)
        and all(column.strip() in result.columns for column in inputs.required_columns)
        and all(list(row) == result.columns for row in result.rows)
    )


READ_EXCEL = Instruction(
    instruction_id=INSTRUCTION_ID,
    name="读取 Excel",
    version="0.1.0",
    description="读取 .xlsx 数据表，第一行为列名；检查必需列，跳过完全空行，不修改原文件，不计算公式。",
    input_model=ReadExcelInput,
    output_model=ReadExcelOutput,
    handler=read_excel,
    verifier=verify_excel,
)


def main() -> int:
    parser = argparse.ArgumentParser(description="读取 Excel 数据表")
    parser.add_argument("file_path", help=".xlsx 文件路径")
    parser.add_argument("--sheet", default=None, help="工作表名称，默认第一张")
    parser.add_argument("--require", nargs="*", default=[], help="必需列名")
    args = parser.parse_args()
    registry = InstructionRegistry()
    registry.register(READ_EXCEL)
    try:
        result = registry.execute(INSTRUCTION_ID, {
            "file_path": args.file_path, "sheet_name": args.sheet, "required_columns": args.require,
        })
    except InstructionError as exc:
        print(json.dumps(exc.to_dict(), ensure_ascii=False), file=sys.stderr)
        return 2
    print(json.dumps(result.model_dump(mode="json"), ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
