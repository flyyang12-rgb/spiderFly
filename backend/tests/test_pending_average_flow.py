from __future__ import annotations

import json
import os
import subprocess
import sys
import tempfile
import unittest
from copy import deepcopy
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from example_flows.excel_pending_average import run_flow
from spiderfly_instructions import InstructionError, InstructionRegistry


class PendingAverageFlowTests(unittest.TestCase):
    def setUp(self) -> None:
        directory = tempfile.TemporaryDirectory(prefix="spiderfly-pending-average-")
        self.addCleanup(directory.cleanup)
        self.directory = Path(directory.name)

    def make_input(self, name: str, columns: list[str], rows: list[list]) -> Path:
        source = self.directory / name
        workbook = Workbook()
        try:
            sheet = workbook.active
            sheet.title = "订单"
            sheet.append(columns)
            for row in rows:
                sheet.append(row)
            sheet["A1"].font = Font(bold=True, color="FF123456")
            sheet["A1"].fill = PatternFill("solid", fgColor="FFEEEECC")
            sheet.column_dimensions["A"].width = 19
            sheet.freeze_panes = "A2"
            if rows:
                sheet["A2"].number_format = "@"
            extra = workbook.create_sheet("说明")
            extra["A1"] = "其他工作表原样保留"
            extra["B2"] = "=SUM(1,2)"
            workbook.save(source)
        finally:
            workbook.close()
        return source

    @staticmethod
    def workbook_values(path: Path) -> dict[str, list[tuple]]:
        workbook = load_workbook(path, data_only=False)
        try:
            return {sheet.title: list(sheet.values) for sheet in workbook.worksheets}
        finally:
            workbook.close()

    def execute_with_trace(self, source: Path, output: Path, sheet_name: str | None = None):
        actual_execute = InstructionRegistry.execute
        calls = []
        successful = []

        def record_execute(registry, instruction_id, inputs=None):
            calls.append((instruction_id, deepcopy(inputs)))
            result = actual_execute(registry, instruction_id, inputs)
            successful.append(instruction_id)
            return result

        with patch.object(InstructionRegistry, "execute", new=record_execute):
            result = run_flow(str(source), str(output), sheet_name)
        return result, calls, successful

    def cli(self, *arguments: str) -> subprocess.CompletedProcess:
        environment = {
            key: value for key, value in os.environ.items()
            if not key.startswith("SPIDERFLY_")
        }
        environment["PYTHONIOENCODING"] = "utf-8"
        return subprocess.run(
            [sys.executable, "-m", "example_flows.excel_pending_average", *arguments],
            cwd=Path(__file__).resolve().parents[1], env=environment,
            capture_output=True, text=True, encoding="utf-8", timeout=15,
        )

    def test_real_instruction_chain_appends_averages_and_preserves_workbook(self) -> None:
        columns = ["订单号", "状态", "金额", "启用", "备注"]
        rows = [
            ["0007", "待处理", "5,2", False, "  原样保留  "],
            ["0088", "已完成", "不能求平均", True, None],
            ["0042", "待处理", 5.6, False, "第三条"],
        ]
        source = self.make_input("input.xlsx", columns, rows)
        output = self.directory / "result.xlsx"
        original = source.read_bytes()
        before = self.workbook_values(source)

        result, calls, successful = self.execute_with_trace(source, output)

        expected_calls = ["excel.read", "math.average", "math.average", "excel.write"]
        self.assertEqual([name for name, _ in calls], expected_calls)
        self.assertEqual(successful, expected_calls)
        self.assertEqual(result, {
            "file_path": str(output.absolute()), "sheet_name": "订单", "row_count": 3,
            "processed_row_count": 2, "instruction_calls": successful,
        })
        self.assertEqual(calls[0][1]["required_columns"], ["状态", "金额"])
        self.assertEqual([inputs for name, inputs in calls if name == "math.average"], [
            {"value": "5,2"}, {"value": 5.6},
        ])
        write_inputs = calls[-1][1]
        self.assertEqual(write_inputs["template_file"], str(source))
        self.assertEqual(write_inputs["columns"], [*columns, "待处理平均数"])
        self.assertEqual(write_inputs["rows"], [
            {**dict(zip(columns, row)), "待处理平均数": average}
            for row, average in zip(rows, [3.5, None, 5.6])
        ])
        self.assertEqual(source.read_bytes(), original)
        after = self.workbook_values(output)
        self.assertEqual(list(after), list(before))
        self.assertEqual(after["说明"], before["说明"])
        self.assertEqual(after["订单"], [
            (*columns, "待处理平均数"),
            (*rows[0], 3.5), (*rows[1], None), (*rows[2], 5.6),
        ])
        workbook = load_workbook(output, data_only=False)
        try:
            sheet = workbook["订单"]
            self.assertEqual(sheet["A2"].value, "0007")
            self.assertEqual(sheet["A2"].data_type, "s")
            self.assertEqual(sheet["C2"].value, "5,2")
            self.assertEqual(sheet["C2"].data_type, "s")
            self.assertEqual(sheet["C4"].value, 5.6)
            self.assertIs(type(sheet["C4"].value), float)
            self.assertIs(sheet["D2"].value, False)
            self.assertIs(sheet["D3"].value, True)
            self.assertEqual(sheet["A2"].number_format, "@")
            self.assertTrue(sheet["A1"].font.bold)
            self.assertEqual(sheet["A1"].fill.fgColor.rgb, "FFEEEECC")
            self.assertEqual(sheet.column_dimensions["A"].width, 19)
            self.assertEqual(sheet.freeze_panes, "A2")
        finally:
            workbook.close()

    def test_empty_table_and_non_pending_bad_amounts_do_not_call_average(self) -> None:
        cases = [
            ("empty", []),
            ("non-pending", [
                ["001", "已完成", "不是数字"],
                ["002", "待处理 ", "5,坏"],
                ["003", " 待处理", True],
                ["004", None, None],
            ]),
        ]
        columns = ["订单号", "状态", "金额"]
        for label, rows in cases:
            with self.subTest(case=label):
                source = self.make_input(f"{label}.xlsx", columns, rows)
                output = self.directory / f"{label}-result.xlsx"
                original = source.read_bytes()
                result, calls, successful = self.execute_with_trace(source, output)
                self.assertEqual([name for name, _ in calls], ["excel.read", "excel.write"])
                self.assertEqual(result["instruction_calls"], successful)
                self.assertEqual(result["processed_row_count"], 0)
                self.assertEqual(result["row_count"], len(rows))
                self.assertEqual(self.workbook_values(output)["订单"], [
                    (*columns, "待处理平均数"), *[(*row, None) for row in rows],
                ])
                self.assertEqual(source.read_bytes(), original)

    def test_bad_pending_amount_stops_before_write_and_keeps_original(self) -> None:
        for label, invalid in (("text", "5,坏"), ("empty", None), ("boolean", True)):
            with self.subTest(case=label):
                source = self.make_input(f"bad-{label}.xlsx", ["状态", "金额"], [
                    ["待处理", "5,2"], ["待处理", invalid],
                ])
                output = self.directory / f"bad-{label}-result.xlsx"
                original = source.read_bytes()
                actual_execute = InstructionRegistry.execute
                calls, successful = [], []

                def record_execute(registry, instruction_id, inputs=None):
                    calls.append(instruction_id)
                    result = actual_execute(registry, instruction_id, inputs)
                    successful.append(instruction_id)
                    return result

                with (
                    patch.object(InstructionRegistry, "execute", new=record_execute),
                    self.assertRaises(InstructionError) as caught,
                ):
                    run_flow(str(source), str(output))
                self.assertEqual(caught.exception.instruction_id, "math.average")
                self.assertEqual(calls, ["excel.read", "math.average", "math.average"])
                self.assertEqual(successful, ["excel.read", "math.average"])
                self.assertFalse(output.exists())
                self.assertEqual(source.read_bytes(), original)

    def test_existing_result_column_is_rejected_before_processing(self) -> None:
        source = self.make_input("existing-column.xlsx", ["状态", "金额", "待处理平均数"], [
            ["待处理", "坏数据", 9],
        ])
        output = self.directory / "existing-column-result.xlsx"
        original = source.read_bytes()
        actual_execute = InstructionRegistry.execute
        calls = []

        def record_execute(registry, instruction_id, inputs=None):
            calls.append(instruction_id)
            return actual_execute(registry, instruction_id, inputs)

        with (
            patch.object(InstructionRegistry, "execute", new=record_execute),
            self.assertRaises(InstructionError) as caught,
        ):
            run_flow(str(source), str(output))
        self.assertEqual(caught.exception.code, "FLOW_COLUMN_EXISTS")
        self.assertEqual(calls, ["excel.read"])
        self.assertFalse(output.exists())
        self.assertEqual(source.read_bytes(), original)

    def test_missing_status_or_amount_fails_without_creating_output(self) -> None:
        for column in ("状态", "金额"):
            with self.subTest(column=column):
                source = self.make_input(f"only-{column}.xlsx", [column], [])
                output = self.directory / f"only-{column}-result.xlsx"
                original = source.read_bytes()
                with self.assertRaises(InstructionError) as caught:
                    run_flow(str(source), str(output))
                self.assertEqual(caught.exception.code, "EXCEL_COLUMNS_MISSING")
                self.assertFalse(output.exists())
                self.assertEqual(source.read_bytes(), original)

    def test_existing_output_and_same_path_are_never_overwritten(self) -> None:
        source = self.make_input("input.xlsx", ["状态", "金额"], [["待处理", 5.6]])
        output = self.directory / "existing.xlsx"
        output.write_bytes(b"existing output must not change")
        original_source, original_output = source.read_bytes(), output.read_bytes()
        for target in (source, output):
            with self.subTest(target=target.name), self.assertRaises(InstructionError) as caught:
                run_flow(str(source), str(target))
            self.assertEqual(caught.exception.code, "EXCEL_FILE_EXISTS")
            self.assertEqual(source.read_bytes(), original_source)
            self.assertEqual(output.read_bytes(), original_output)

    def test_cli_selects_sheet_and_returns_json_success_or_instruction_error(self) -> None:
        source = self.make_input("multiple-sheets.xlsx", ["状态", "金额"], [["待处理", "5,2"]])
        workbook = load_workbook(source)
        try:
            workbook.move_sheet("说明", offset=-1)
            workbook.save(source)
        finally:
            workbook.close()
        original = source.read_bytes()
        output = self.directory / "cli-result.xlsx"
        completed = self.cli(str(source), str(output), "--sheet", "订单")
        self.assertEqual(completed.returncode, 0, completed.stderr)
        self.assertEqual(completed.stderr, "")
        result = json.loads(completed.stdout)
        self.assertEqual(result["sheet_name"], "订单")
        self.assertEqual(result["row_count"], 1)
        self.assertEqual(result["processed_row_count"], 1)
        self.assertEqual(result["instruction_calls"], ["excel.read", "math.average", "excel.write"])
        self.assertEqual(self.workbook_values(output)["订单"], [
            ("状态", "金额", "待处理平均数"), ("待处理", "5,2", 3.5),
        ])
        failed_output = self.directory / "failed-cli-result.xlsx"
        failed = self.cli(str(source), str(failed_output), "--sheet", "不存在")
        self.assertNotEqual(failed.returncode, 0)
        self.assertEqual(failed.stdout, "")
        self.assertEqual(json.loads(failed.stderr)["code"], "EXCEL_SHEET_MISSING")
        self.assertFalse(failed_output.exists())
        self.assertEqual(source.read_bytes(), original)


if __name__ == "__main__":
    unittest.main()
