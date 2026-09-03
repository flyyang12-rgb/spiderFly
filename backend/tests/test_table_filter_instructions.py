from __future__ import annotations

import tempfile
import unittest
from copy import deepcopy
from dataclasses import replace
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook

from spiderfly_instructions import InstructionError, InstructionRegistry
from spiderfly_instructions.excel import READ_EXCEL
from spiderfly_instructions.excel_write import WRITE_EXCEL
from spiderfly_instructions.table_filter import FILTER_EQUALS


class TableFilterInstructionTests(unittest.TestCase):
    def setUp(self) -> None:
        self.registry = InstructionRegistry()
        self.registry.register(FILTER_EQUALS)

    def parameters(self, **changes) -> dict:
        values = {
            "columns": ["编号", "状态"],
            "rows": [{"编号": "001", "状态": "待处理"}],
            "column": "状态",
            "value": "待处理",
        }
        values.update(changes)
        return values

    def execute(self, **changes):
        return self.registry.execute("table.filter_equals", self.parameters(**changes))

    def assert_filter_error(self, expected: str | tuple[str, ...], **changes) -> None:
        with self.assertRaises(InstructionError) as caught:
            self.execute(**changes)
        allowed = (expected,) if isinstance(expected, str) else expected
        self.assertIn(caught.exception.code, allowed)
        self.assertEqual(caught.exception.instruction_id, "table.filter_equals")

    def test_exact_text_preserves_columns_row_order_duplicates_and_input(self) -> None:
        columns = ["编号", "状态", "备注"]
        first = {"备注": "  原样保留  ", "状态": "Ready", "编号": "0007"}
        last = {"编号": "0002", "状态": "Ready", "备注": None}
        rows = [
            first,
            {"编号": "0010", "状态": "ready", "备注": "小写不同"},
            {"编号": "0011", "状态": " Ready", "备注": "前空格不同"},
            {"编号": "0012", "状态": "Ready ", "备注": "后空格不同"},
            deepcopy(first),
            last,
        ]
        parameters = self.parameters(columns=columns, rows=rows, value="Ready")
        original = deepcopy(parameters)

        result = self.registry.execute("table.filter_equals", parameters)

        self.assertEqual(result.model_dump(), {
            "columns": columns, "rows": [first, first, last], "row_count": 3,
        })
        self.assertTrue(all(list(row) == columns for row in result.rows))
        self.assertEqual(parameters, original)
        result.rows[0]["备注"] = "只改返回结果"
        result.columns.append("返回列")
        self.assertEqual(parameters, original)
        self.assertEqual(result.rows[1]["备注"], "  原样保留  ")

    def test_numbers_booleans_text_and_empty_values_remain_distinct(self) -> None:
        values = [1, 1.0, True, "1", 0, 0.0, False, None, "", " "]
        rows = [{"编号": str(index), "状态": value} for index, value in enumerate(values)]
        cases = [
            (1, [0, 1]), (1.0, [0, 1]), (True, [2]), ("1", [3]),
            (0, [4, 5]), (False, [6]), (None, [7]), ("", [8]), (" ", [9]),
        ]
        for value, expected_indices in cases:
            with self.subTest(value=value, type=type(value).__name__):
                result = self.execute(rows=rows, value=value)
                self.assertEqual(result.rows, [rows[index] for index in expected_indices])
                self.assertEqual(result.row_count, len(expected_indices))
                for actual, index in zip(result.rows, expected_indices):
                    self.assertIs(type(actual["状态"]), type(values[index]))

    def test_excel_cell_types_are_preserved_and_can_be_compared(self) -> None:
        cell_values = {
            "日期": date(2026, 9, 10),
            "时间点": datetime(2026, 9, 10, 8, 30, 1),
            "时刻": time(8, 30, 1),
            "时长": timedelta(days=1, milliseconds=125),
            "带时区时间点": datetime(2026, 9, 10, 8, 30, tzinfo=timezone.utc),
            "带时区时刻": time(8, 30, tzinfo=timezone.utc),
            "启用": False,
            "数量": 0,
            "金额": 2.5,
            "空值": None,
            "空文字": "",
            "原文": "=原样文字",
        }
        columns = ["编号", "状态", *cell_values]
        row = {"编号": "00123", "状态": "待处理", **cell_values}
        result = self.execute(columns=columns, rows=[row])
        self.assertEqual(result.columns, columns)
        self.assertEqual(result.rows, [row])
        for field, value in cell_values.items():
            self.assertIs(type(result.rows[0][field]), type(value))

        temporal_values = list(cell_values.values())[:4]
        temporal_rows = [{"编号": str(index), "状态": value} for index, value in enumerate(temporal_values)]
        temporal_rows.append({"编号": "date-text", "状态": "2026-09-10"})
        for index, value in enumerate(temporal_values):
            with self.subTest(type=type(value).__name__):
                filtered = self.execute(rows=temporal_rows, value=value)
                self.assertEqual(filtered.rows, [temporal_rows[index]])
                self.assertIs(type(filtered.rows[0]["状态"]), type(value))

    def test_empty_tables_and_no_matches_return_the_original_columns(self) -> None:
        for rows in ([], [{"编号": "0001", "状态": "已处理"}]):
            with self.subTest(rows=rows):
                self.assertEqual(self.execute(rows=rows).model_dump(), {
                    "columns": ["编号", "状态"], "rows": [], "row_count": 0,
                })

    def test_all_parameters_are_required_and_wrong_types_are_rejected(self) -> None:
        for field in self.parameters():
            with self.subTest(missing=field):
                parameters = self.parameters()
                parameters.pop(field)
                with self.assertRaises(InstructionError) as caught:
                    self.registry.execute("table.filter_equals", parameters)
                self.assertEqual(caught.exception.code, "INPUT_INVALID")
        for changes in (
            {"columns": "状态"}, {"rows": {}}, {"column": 1},
            {"value": ["待处理"]}, {"value": {"状态": "待处理"}},
        ):
            with self.subTest(changes=changes):
                self.assert_filter_error("INPUT_INVALID", **changes)
        self.assert_filter_error(("INPUT_INVALID", "TABLE_COLUMNS_INVALID"), columns=[], rows=[])
        self.assert_filter_error(("INPUT_INVALID", "TABLE_COLUMN_MISSING"), column="")

    def test_invalid_column_names_are_checked_even_when_there_are_no_rows(self) -> None:
        for columns in ([""], [" "], [" 状态"], ["状态 "], ["状态", "状态"]):
            with self.subTest(columns=columns):
                self.assert_filter_error("TABLE_COLUMNS_INVALID", columns=columns, rows=[])

    def test_missing_filter_column_and_malformed_unmatched_rows_are_rejected(self) -> None:
        self.assert_filter_error("TABLE_COLUMN_MISSING", columns=["编号"], rows=[])
        self.assert_filter_error("TABLE_COLUMN_MISSING", column="状态 ")
        for row in (
            {"编号": "0002"},
            {"状态": "不匹配"},
            {"编号": "0002", "状态": "不匹配", "额外": "不能静默丢弃"},
        ):
            with self.subTest(row=row):
                self.assert_filter_error("TABLE_ROW_INVALID", rows=[{"编号": "0001", "状态": "待处理"}, row])

    def test_nonfinite_filter_values_and_cells_fail_instead_of_silently_matching(self) -> None:
        for value in (float("nan"), float("inf"), float("-inf")):
            with self.subTest(location="filter-value", value=value):
                self.assert_filter_error("TABLE_VALUE_INVALID", rows=[], value=value)
            with self.subTest(location="filter-column", value=value):
                self.assert_filter_error("TABLE_VALUE_INVALID", rows=[
                    {"编号": "0001", "状态": "待处理"}, {"编号": "0002", "状态": value},
                ])

    def test_timezone_aware_filter_values_and_cells_require_explicit_conversion(self) -> None:
        for value in (
            datetime(2026, 9, 10, 8, 30, tzinfo=timezone.utc),
            time(8, 30, tzinfo=timezone(timedelta(hours=8))),
        ):
            with self.subTest(location="filter-value", type=type(value).__name__):
                self.assert_filter_error("TABLE_VALUE_INVALID", rows=[], value=value)
            with self.subTest(location="filter-column", type=type(value).__name__):
                self.assert_filter_error("TABLE_VALUE_INVALID", rows=[{"编号": "0001", "状态": value}])

    def test_verifier_rejects_wrong_rows_even_with_correct_count_and_filter_values(self) -> None:
        def wrong_rows(inputs):
            return {
                "columns": inputs.columns,
                "rows": [inputs.rows[0], inputs.rows[0]],
                "row_count": 2,
            }

        registry = InstructionRegistry()
        registry.register(replace(FILTER_EQUALS, handler=wrong_rows))
        with self.assertRaises(InstructionError) as caught:
            registry.execute("table.filter_equals", self.parameters(rows=[
                {"编号": "0001", "状态": "待处理"},
                {"编号": "0002", "状态": "已处理"},
                {"编号": "0003", "状态": "待处理"},
            ]))
        self.assertEqual(caught.exception.code, "VERIFICATION_FAILED")
        self.assertEqual(caught.exception.stage, "verify")

    def test_verifier_rejects_numeric_values_replaced_by_equal_booleans(self) -> None:
        def changed_type(inputs):
            return {
                "columns": inputs.columns,
                "rows": [{**inputs.rows[0], "数量": False}],
                "row_count": 1,
            }

        registry = InstructionRegistry()
        registry.register(replace(FILTER_EQUALS, handler=changed_type))
        with self.assertRaises(InstructionError) as caught:
            registry.execute("table.filter_equals", self.parameters(
                columns=["状态", "数量"], rows=[{"状态": "待处理", "数量": 0}],
            ))
        self.assertEqual(caught.exception.code, "VERIFICATION_FAILED")

    def test_verifier_rejects_equivalent_timezone_conversion_in_an_unfiltered_column(self) -> None:
        local_zone = timezone(timedelta(hours=8))
        cases = (
            (datetime(2026, 9, 10, 9, tzinfo=local_zone), datetime(2026, 9, 10, 1, tzinfo=timezone.utc)),
            (time(9, tzinfo=local_zone), time(1, tzinfo=timezone.utc)),
        )
        for original, changed in cases:
            with self.subTest(type=type(original).__name__):
                self.assertEqual(original, changed)

                def changed_timezone(inputs):
                    return {
                        "columns": inputs.columns,
                        "rows": [{**inputs.rows[0], "时间": changed}],
                        "row_count": 1,
                    }

                registry = InstructionRegistry()
                registry.register(replace(FILTER_EQUALS, handler=changed_timezone))
                with self.assertRaises(InstructionError) as caught:
                    registry.execute("table.filter_equals", self.parameters(
                        columns=["状态", "时间"], rows=[{"状态": "待处理", "时间": original}],
                    ))
                self.assertEqual(caught.exception.code, "VERIFICATION_FAILED")

    def test_read_filter_write_excel_preserves_source_and_expected_records(self) -> None:
        with tempfile.TemporaryDirectory(prefix="spiderfly-filter-excel-") as directory:
            source = Path(directory) / "订单.xlsx"
            output = Path(directory) / "待处理.xlsx"
            columns = ["订单号", "状态", "数量", "启用", "备注"]
            workbook = Workbook()
            try:
                sheet = workbook.active
                sheet.title = "订单"
                sheet.append(columns)
                sheet.append(["00123", "待处理", 0, False, "  待核对  "])
                sheet.append(["00999", "已处理", 4, True, "不保留"])
                sheet.append(["00001", "待处理", 2, True, "保留"])
                workbook.save(source)
            finally:
                workbook.close()
            original = source.read_bytes()
            self.registry.register(READ_EXCEL)
            self.registry.register(WRITE_EXCEL)

            table = self.registry.execute("excel.read", {"file_path": str(source)})
            filtered = self.registry.execute("table.filter_equals", {
                "columns": table.columns, "rows": table.rows, "column": "状态", "value": "待处理",
            })
            saved = self.registry.execute("excel.write", {
                "file_path": str(output), "sheet_name": table.sheet_name,
                "columns": filtered.columns, "rows": filtered.rows,
            })

            self.assertEqual(saved.row_count, 2)
            self.assertEqual(source.read_bytes(), original)
            written = load_workbook(output, read_only=True, data_only=False)
            try:
                sheet = written["订单"]
                self.assertEqual(list(sheet.values), [
                    tuple(columns),
                    ("00123", "待处理", 0, False, "  待核对  "),
                    ("00001", "待处理", 2, True, "保留"),
                ])
                self.assertEqual(sheet["A2"].data_type, "s")
                self.assertIs(type(sheet["C2"].value), int)
                self.assertIs(sheet["D2"].value, False)
            finally:
                written.close()


if __name__ == "__main__":
    unittest.main()
