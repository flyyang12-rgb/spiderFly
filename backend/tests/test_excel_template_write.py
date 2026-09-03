from __future__ import annotations

import tempfile
import unittest
from copy import deepcopy
from dataclasses import replace
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook

from spiderfly_instructions import InstructionError, InstructionRegistry
from spiderfly_instructions.excel import READ_EXCEL
from spiderfly_instructions.excel_write import WRITE_EXCEL


class ExcelTemplateWriteTests(unittest.TestCase):
    def setUp(self) -> None:
        directory = tempfile.TemporaryDirectory(prefix="spiderfly-template-write-")
        self.addCleanup(directory.cleanup)
        self.directory = Path(directory.name)
        self.source = self.directory / "原工作簿.xlsx"
        self.output = self.directory / "追加结果.xlsx"
        workbook = Workbook()
        try:
            sheet = workbook.active
            sheet.title = "订单"
            sheet.append(["编号", "金额", "启用", "日期"])
            sheet.append(["0001", 0, False, date(2026, 9, 3)])
            sheet.append([None, None, None, None])
            sheet.append(["0002", 5.6, True, date(2026, 9, 4)])
            sheet.row_dimensions[3].height = 27
            extra = workbook.create_sheet("说明")
            extra["A1"] = "保留其他页"
            extra["C2"] = "=1+2"
            extra["B8"] = "末行也不能丢失"
            workbook.save(self.source)
        finally:
            workbook.close()
        self.original = self.source.read_bytes()
        self.registry = InstructionRegistry()
        self.registry.register(READ_EXCEL)
        self.registry.register(WRITE_EXCEL)
        self.table = self.registry.execute("excel.read", {"file_path": str(self.source)})

    def parameters(self, **changes) -> dict:
        values = {
            "file_path": str(self.output),
            "template_file": str(self.source),
            "sheet_name": self.table.sheet_name,
            "columns": [*self.table.columns, "结果"],
            "rows": [
                {**row, "结果": result} for row, result in zip(self.table.rows, [7 / 3, 0])
            ],
        }
        values.update(changes)
        return values

    def assert_rejected(self, parameters: dict, code: str) -> None:
        with self.assertRaises(InstructionError) as caught:
            self.registry.execute("excel.write", parameters)
        self.assertEqual(caught.exception.instruction_id, "excel.write")
        self.assertEqual(caught.exception.code, code)
        self.assertFalse(Path(parameters["file_path"]).exists())
        self.assertEqual(self.source.read_bytes(), self.original)

    def test_read_skips_blank_row_but_append_preserves_physical_positions_and_original_dates(self) -> None:
        self.assertEqual(self.table.row_count, 2)
        parameters = self.parameters()
        original_parameters = deepcopy(parameters)
        result = self.registry.execute("excel.write", parameters)
        self.assertEqual(result.row_count, 2)
        self.assertEqual(parameters, original_parameters)
        before, after = load_workbook(self.source), load_workbook(self.output)
        try:
            self.assertEqual(after.sheetnames, before.sheetnames)
            self.assertEqual(list(after["说明"].values), list(before["说明"].values))
            original_sheet, sheet = before["订单"], after["订单"]
            self.assertEqual(sheet.max_row, 4)
            self.assertEqual(sheet.max_column, 5)
            self.assertEqual(tuple(cell.value for cell in sheet[1]), (*self.table.columns, "结果"))
            self.assertEqual(tuple(cell.value for cell in sheet[3]), (None,) * 5)
            self.assertEqual(sheet.row_dimensions[3].height, 27)
            self.assertAlmostEqual(sheet["E2"].value, 7 / 3, places=14)
            self.assertEqual(sheet["E4"].value, 0)
            self.assertIs(type(sheet["E4"].value), int)
            for row in (2, 4):
                for column in range(1, 5):
                    previous, written = original_sheet.cell(row, column), sheet.cell(row, column)
                    self.assertEqual(written.value, previous.value)
                    self.assertIs(type(written.value), type(previous.value))
                    self.assertEqual(written.data_type, previous.data_type)
            self.assertEqual(sheet["A2"].value, "0001")
            self.assertIs(sheet["C2"].value, False)
        finally:
            before.close()
            after.close()
        self.assertEqual(self.source.read_bytes(), self.original)

    def test_wps_outline_summary_is_recomputed_without_changing_actual_dimensions(self) -> None:
        # WPS can retain a maximum group level even when no column has that level.
        with ZipFile(self.source) as archive:
            entries = [(item, archive.read(item.filename)) for item in archive.infolist()]
        buffer = BytesIO()
        with ZipFile(buffer, "w") as archive:
            for item, data in entries:
                if item.filename == "xl/worksheets/sheet1.xml":
                    root = ElementTree.fromstring(data)
                    root.find("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheetFormatPr").set("outlineLevelCol", "2")
                    data = ElementTree.tostring(root, encoding="utf-8")
                archive.writestr(item, data)
        self.source.write_bytes(buffer.getvalue())
        original = self.source.read_bytes()
        result = self.registry.execute("excel.write", self.parameters())
        self.assertEqual(result.row_count, 2)
        self.assertEqual(self.source.read_bytes(), original)
        workbook = load_workbook(self.output)
        try:
            self.assertAlmostEqual(workbook["订单"]["E2"].value, 7 / 3, places=14)
            self.assertEqual(workbook["订单"].row_dimensions[3].height, 27)
            self.assertEqual(workbook["订单"].column_dimensions["A"].outlineLevel, 0)
        finally:
            workbook.close()

    def test_changed_original_data_missing_rows_reordered_columns_or_no_addition_are_rejected(self) -> None:
        altered = self.parameters()
        altered["rows"][0]["金额"] = 99
        boolean = self.parameters()
        boolean["rows"][0]["金额"] = False
        missing = self.parameters(rows=self.parameters()["rows"][:-1])
        reordered_rows = self.parameters(rows=list(reversed(self.parameters()["rows"])))
        reordered_columns = self.parameters(columns=["金额", "编号", "启用", "日期", "结果"])
        no_addition = self.parameters(columns=list(self.table.columns), rows=deepcopy(self.table.rows))
        for label, parameters in (
            ("changed-value", altered), ("zero-to-false", boolean), ("missing-row", missing),
            ("reordered-rows", reordered_rows), ("reordered-columns", reordered_columns),
            ("no-new-column", no_addition),
        ):
            with self.subTest(case=label):
                self.assert_rejected(parameters, "EXCEL_TEMPLATE_MISMATCH")

    def test_existing_output_and_template_as_output_are_not_overwritten(self) -> None:
        self.output.write_bytes(b"do not overwrite this file")
        previous_output = self.output.read_bytes()
        for target in (self.output, self.source):
            with self.subTest(target=target.name), self.assertRaises(InstructionError) as caught:
                self.registry.execute("excel.write", self.parameters(file_path=str(target)))
            self.assertEqual(caught.exception.code, "EXCEL_FILE_EXISTS")
            self.assertEqual(self.output.read_bytes(), previous_output)
            self.assertEqual(self.source.read_bytes(), self.original)

    def test_appended_cells_support_public_values_and_reject_unrepresentable_values(self) -> None:
        additions = {
            "文本编号": "00123", "文字原文": "=1+1", "日期结果": date(2026, 9, 5),
            "时间点": datetime(2026, 9, 5, 8, 20, 1, 123000),
            "时刻": time(8, 20, 1, 123000), "时长": timedelta(days=1, milliseconds=123),
            "整数": 0, "布尔": False, "空值": None, "空文字": "",
        }
        parameters = self.parameters(
            columns=[*self.table.columns, *additions],
            rows=[{**row, **additions} for row in self.table.rows],
        )
        result = self.registry.execute("excel.write", parameters)
        table = self.registry.execute("excel.read", {"file_path": result.file_path})
        for index, row in enumerate(table.rows):
            self.assertEqual(row, {**self.table.rows[index], **additions, "空文字": None})
            for name, expected in additions.items():
                if name != "空文字":
                    self.assertIs(type(row[name]), type(expected))
        workbook = load_workbook(self.output)
        try:
            self.assertEqual(workbook["订单"]["F2"].data_type, "s")
            self.assertEqual(workbook["订单"]["F2"].value, "=1+1")
        finally:
            workbook.close()
        for label, value in (
            ("nonfinite", float("nan")),
            ("timezone", datetime(2026, 9, 5, tzinfo=timezone.utc)),
            ("sub-millisecond", time(8, 20, microsecond=1)),
        ):
            with self.subTest(case=label):
                invalid = self.parameters(file_path=str(self.directory / f"{label}.xlsx"))
                invalid["rows"][0]["结果"] = value
                self.assert_rejected(invalid, "EXCEL_VALUE_INVALID")
        self.assertEqual(self.source.read_bytes(), self.original)

    def test_verifier_rejects_corrupted_originals_additions_removed_data_and_missing_sheets(self) -> None:
        def change_original(workbook):
            workbook["订单"]["B2"] = 99

        def change_original_type(workbook):
            workbook["订单"]["B2"] = False

        def change_result(workbook):
            workbook["订单"]["E2"] = 5

        def false_result(workbook):
            workbook["订单"]["E2"] = False

        def zero_result_as_false(workbook):
            workbook["订单"]["E4"] = False

        def remove_original_row(workbook):
            workbook["订单"].delete_rows(4)

        def remove_other_page_tail(workbook):
            workbook["说明"].delete_rows(8)

        def remove_sheet(workbook):
            del workbook["说明"]

        def remove_results(workbook):
            workbook["订单"].delete_cols(5)

        def move_result_to_blank_row(workbook):
            workbook["订单"]["E3"] = workbook["订单"]["E4"].value
            workbook["订单"]["E4"] = None

        def change_original_column_width(workbook):
            workbook["订单"].column_dimensions["A"].width = 99

        def hide_other_sheet(workbook):
            workbook["说明"].sheet_state = "hidden"

        def change_column_group(workbook):
            workbook["订单"].column_dimensions["A"].outlineLevel = 2

        def change_row_group(workbook):
            workbook["订单"].row_dimensions[4].outlineLevel = 1

        for mutate in (
            change_original, change_original_type, change_result, false_result,
            zero_result_as_false, remove_original_row, remove_other_page_tail,
            remove_sheet, remove_results, move_result_to_blank_row,
            change_original_column_width, hide_other_sheet, change_column_group, change_row_group,
        ):
            with self.subTest(case=mutate.__name__):
                output = self.directory / f"{mutate.__name__}.xlsx"

                def corrupt_after_write(inputs):
                    result = WRITE_EXCEL.handler(inputs)
                    workbook = load_workbook(result["file_path"])
                    try:
                        mutate(workbook)
                        workbook.save(result["file_path"])
                    finally:
                        workbook.close()
                    return result

                registry = InstructionRegistry()
                registry.register(replace(WRITE_EXCEL, handler=corrupt_after_write))
                with self.assertRaises(InstructionError) as caught:
                    registry.execute("excel.write", self.parameters(file_path=str(output)))
                self.assertEqual(caught.exception.code, "VERIFICATION_FAILED")
                self.assertEqual(caught.exception.instruction_id, "excel.write")
                self.assertEqual(self.source.read_bytes(), self.original)


if __name__ == "__main__":
    unittest.main()
