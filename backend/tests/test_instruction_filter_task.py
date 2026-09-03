from __future__ import annotations

import importlib.util
import io
import json
import os
import shutil
import subprocess
import sys
import tempfile
import unittest
from contextlib import redirect_stderr, redirect_stdout
from datetime import datetime
from hashlib import sha256
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook


BACKEND = Path(__file__).resolve().parents[1]
EXAMPLE = BACKEND.parent / "examples" / "instruction_excel_filter_task.py"


class InstructionFilterTaskTests(unittest.TestCase):
    def setUp(self) -> None:
        directory = tempfile.TemporaryDirectory(prefix="spiderfly-filter-task-")
        self.addCleanup(directory.cleanup)
        self.directory = Path(directory.name)
        self.working_dir = self.directory / "working"
        self.working_dir.mkdir()
        self.make_execution("execution")
        self.columns = ["金额", "订单号", "状态", "启用", "备注", "登记时间"]
        self.input_rows = [
            [0, "0007", "待处理", False, "  原样保留  ", datetime(2026, 9, 3, 8, 20)],
            [2.5, "0099", "已完成", True, "不输出", None],
            [8.75, "0011", "待处理 ", True, "末尾空格不匹配", None],
            [125, "0420", "待处理", True, None, None],
        ]
        self.upload_file = self.make_input("上传订单.xlsx", self.columns, self.input_rows)

    def make_execution(self, name: str) -> None:
        self.artifact_dir = self.directory / name / "artifacts"
        self.artifact_dir.mkdir(parents=True)
        self.result_file = self.artifact_dir.parent / "result.json"
        self.output_dir = self.artifact_dir / "订单筛选示例"

    def make_input(self, name: str, columns: list[str], rows: list[list]) -> Path:
        path = self.directory / name
        workbook = Workbook()
        try:
            workbook.active.title = "上传订单"
            workbook.active.append(columns)
            for row in rows:
                workbook.active.append(row)
            # 第二张表不能被误选，也不应出现在结果文件里。
            other = workbook.create_sheet("其他表")
            other.append(["订单号", "状态"])
            other.append(["不可混入", "待处理"])
            workbook.save(path)
        finally:
            workbook.close()
        return path

    def environment(self) -> dict[str, str]:
        environment = {
            key: value for key, value in os.environ.items()
            if not key.upper().startswith(("SPIDERFLY_", "FEISHU_"))
            and key.upper() not in ("PYTHONPATH", "PYTHONHOME")
        }
        environment.update({
            "PYTHONIOENCODING": "utf-8",
            "PYTHONPATH": str(BACKEND),
            "SPIDERFLY_RESULT_FILE": str(self.result_file),
            "SPIDERFLY_ARTIFACT_DIR": str(self.artifact_dir),
            "SPIDERFLY_TEMPLATE_FILE": str(self.upload_file),
        })
        return environment

    def run_example(
        self, *, missing: str | None = None, without_package: bool = False,
        template_path: Path | None = None, package_path: Path | None = None,
    ) -> subprocess.CompletedProcess:
        environment = self.environment()
        if template_path is not None:
            environment["SPIDERFLY_TEMPLATE_FILE"] = str(template_path)
        if missing is not None:
            environment.pop(missing)
        if package_path is not None:
            environment["PYTHONPATH"] = str(package_path)
        arguments = [sys.executable]
        if without_package:
            arguments.append("-S")
            environment["PYTHONPATH"] = ""
        return subprocess.run(
            [*arguments, str(EXAMPLE)], cwd=self.working_dir, env=environment,
            capture_output=True, text=True, encoding="utf-8", timeout=15,
        )

    @staticmethod
    def load_example():
        spec = importlib.util.spec_from_file_location("instruction_filter_task_test", EXAMPLE)
        example = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(example)
        return example

    def run_main(self, example) -> subprocess.CompletedProcess:
        stdout, stderr = io.StringIO(), io.StringIO()
        with (
            patch.dict(os.environ, self.environment(), clear=True),
            redirect_stdout(stdout), redirect_stderr(stderr),
        ):
            exit_code = example.main()
        return subprocess.CompletedProcess([], exit_code, stdout.getvalue(), stderr.getvalue())

    def receipt(self) -> dict:
        return json.loads(self.result_file.read_text(encoding="utf-8"))

    def output_rows(self) -> list[tuple]:
        workbook = load_workbook(self.output_dir / "结果.xlsx", read_only=True)
        try:
            self.assertEqual(workbook.sheetnames, ["上传订单"])
            return list(workbook.active.values)
        finally:
            workbook.close()

    def assert_receipt(self, outcome: str, code: str) -> dict:
        receipt = self.receipt()
        self.assertEqual(receipt["schema_version"], 1)
        self.assertEqual(receipt["outcome"], outcome)
        self.assertEqual(receipt["code"], code)
        self.assertIs(receipt["retryable"], False)
        self.assertLessEqual(len(receipt["message"]), 1000)
        self.assertTrue(receipt["message"])
        self.assertEqual(list(self.working_dir.iterdir()), [])
        self.assertEqual(list(self.result_file.parent.glob(".result-*.tmp")), [])
        return receipt

    def assert_failed(self, completed: subprocess.CompletedProcess) -> None:
        self.assertNotEqual(completed.returncode, 0)
        self.assert_receipt("failure", "EXCEL_FILTER_FAILED")
        self.assertTrue(completed.stderr)
        self.assertNotIn("保留", completed.stdout)

    def assert_input_unchanged(self, source: Path, original: bytes) -> None:
        expected_hash = sha256(original).hexdigest()
        self.assertEqual(sha256(source.read_bytes()).hexdigest(), expected_hash)
        self.assertEqual(
            sha256((self.output_dir / "输入.xlsx").read_bytes()).hexdigest(), expected_hash,
        )

    def test_uploaded_table_filters_first_sheet_and_preserves_data_and_receipt(self) -> None:
        original = self.upload_file.read_bytes()
        completed = self.run_example()
        self.assertEqual(completed.returncode, 0, completed.stderr)
        self.assertEqual(completed.stderr, "")
        rows = self.output_rows()
        self.assertEqual(rows, [tuple(self.columns), tuple(self.input_rows[0]), tuple(self.input_rows[3])])
        self.assertIs(type(rows[1][0]), int)
        self.assertIs(rows[1][3], False)
        self.assertIs(rows[2][3], True)
        self.assertIs(type(rows[1][5]), datetime)
        self.assert_input_unchanged(self.upload_file, original)
        receipt = self.assert_receipt("success", "EXCEL_FILTER_DONE")
        self.assertIn("读取 4 行，保留 2 行", receipt["message"])
        self.assertIn(str(self.output_dir / "结果.xlsx"), receipt["message"])
        self.assertIn(receipt["message"], completed.stdout)
        self.assertEqual({path.name for path in self.output_dir.iterdir()}, {"输入.xlsx", "结果.xlsx"})

    def test_no_matches_and_empty_table_succeed_with_original_headers(self) -> None:
        for label, rows in (("no-matches", [["001", "已完成"], ["002", None]]), ("empty", [])):
            with self.subTest(case=label):
                self.make_execution(label)
                source = self.make_input(f"{label}.xlsx", ["订单号", "状态"], rows)
                original = source.read_bytes()
                completed = self.run_example(template_path=source)
                self.assertEqual(completed.returncode, 0, completed.stderr)
                self.assertEqual(self.output_rows(), [("订单号", "状态")])
                receipt = self.assert_receipt("success", "EXCEL_FILTER_DONE")
                self.assertIn(f"读取 {len(rows)} 行，保留 0 行", receipt["message"])
                self.assert_input_unchanged(source, original)

    def test_missing_status_column_keeps_input_and_produces_no_result(self) -> None:
        for label, rows in (("with-rows", [["001", 7]]), ("empty", [])):
            with self.subTest(case=label):
                self.make_execution(f"missing-column-{label}")
                source = self.make_input(f"missing-{label}.xlsx", ["订单号", "金额"], rows)
                original = source.read_bytes()
                completed = self.run_example(template_path=source)
                self.assert_failed(completed)
                self.assertIn("缺少必需列", completed.stderr)
                self.assertIn("状态", completed.stderr)
                self.assert_input_unchanged(source, original)
                self.assertEqual({path.name for path in self.output_dir.iterdir()}, {"输入.xlsx"})

    def test_missing_upload_fails_without_creating_sample_data(self) -> None:
        completed = self.run_example(missing="SPIDERFLY_TEMPLATE_FILE")
        self.assert_failed(completed)
        self.assertIn("上传", completed.stderr)
        self.assertIn("Excel 模板", completed.stderr)
        self.assertEqual(list(self.artifact_dir.iterdir()), [])

    def test_missing_file_wrong_format_directory_and_corrupt_workbook_fail(self) -> None:
        wrong_format = self.directory / "订单.csv"
        wrong_format.write_bytes(self.upload_file.read_bytes())
        directory_input = self.directory / "目录.xlsx"
        directory_input.mkdir()
        corrupt = self.directory / "损坏.xlsx"
        corrupt.write_bytes(b"This is not an Excel workbook")
        for label, source in (
            ("missing", self.directory / "不存在.xlsx"), ("wrong-format", wrong_format),
            ("directory", directory_input), ("corrupt", corrupt),
        ):
            with self.subTest(case=label):
                self.make_execution(label)
                original = source.read_bytes() if source.is_file() else None
                completed = self.run_example(template_path=source)
                self.assert_failed(completed)
                self.assertFalse((self.output_dir / "结果.xlsx").exists())
                if label == "corrupt":
                    self.assert_input_unchanged(source, original)
                else:
                    self.assertEqual(list(self.artifact_dir.iterdir()), [])
                if original is not None:
                    self.assertEqual(source.read_bytes(), original)

    def test_missing_platform_paths_never_write_to_working_directory(self) -> None:
        for variable in ("SPIDERFLY_RESULT_FILE", "SPIDERFLY_ARTIFACT_DIR"):
            with self.subTest(variable=variable):
                self.make_execution(variable)
                completed = self.run_example(missing=variable)
                self.assertNotEqual(completed.returncode, 0)
                self.assertIn(variable, completed.stderr)
                if variable == "SPIDERFLY_RESULT_FILE":
                    self.assertFalse(self.result_file.exists())
                else:
                    self.assert_failed(completed)
                self.assertEqual(list(self.artifact_dir.iterdir()), [])
                self.assertEqual(list(self.working_dir.iterdir()), [])

    def test_missing_instruction_package_has_actionable_failure(self) -> None:
        original = self.upload_file.read_bytes()
        completed = self.run_example(without_package=True)
        self.assert_failed(completed)
        self.assertIn("spiderfly-instructions==0.1.1", completed.stderr)
        self.assert_input_unchanged(self.upload_file, original)
        self.assertFalse((self.output_dir / "结果.xlsx").exists())

    def test_old_package_without_filter_has_actionable_failure(self) -> None:
        # 模拟仍只有旧三条指令的环境；兼容源码及已安装 wheel 两种测试方式。
        package_spec = importlib.util.find_spec("spiderfly_instructions")
        package_dir = Path(package_spec.origin).parent
        old_package = self.directory / "old-package"
        shutil.copytree(
            package_dir, old_package / "spiderfly_instructions",
            ignore=shutil.ignore_patterns("table_filter.py", "__pycache__"),
        )
        original = self.upload_file.read_bytes()
        completed = self.run_example(package_path=old_package)
        self.assert_failed(completed)
        self.assertIn("spiderfly-instructions==0.1.1", completed.stderr)
        self.assert_input_unchanged(self.upload_file, original)
        self.assertFalse((self.output_dir / "结果.xlsx").exists())

    def test_repeated_execution_fails_without_overwriting_artifacts(self) -> None:
        first = self.run_example()
        self.assertEqual(first.returncode, 0, first.stderr)
        before = {path: path.read_bytes() for path in self.artifact_dir.rglob("*") if path.is_file()}
        original = self.upload_file.read_bytes()
        repeated = self.run_example()
        self.assert_failed(repeated)
        self.assertIn("新目录", repeated.stderr)
        self.assertEqual(
            {path: path.read_bytes() for path in self.artifact_dir.rglob("*") if path.is_file()}, before,
        )
        self.assertEqual(self.upload_file.read_bytes(), original)

    def test_interrupted_copy_cleans_partial_input_and_never_runs_filter(self) -> None:
        example = self.load_example()
        original = self.upload_file.read_bytes()

        def interrupted_copy(incoming, outgoing) -> None:
            outgoing.write(incoming.read(16))
            outgoing.flush()
            raise OSError("模拟输入复制中断")

        with (
            patch.object(example.shutil, "copyfileobj", side_effect=interrupted_copy) as copy,
            patch.object(example, "run_filter") as run_filter,
        ):
            completed = self.run_main(example)
        copy.assert_called_once()
        run_filter.assert_not_called()
        self.assert_failed(completed)
        self.assertIn("模拟输入复制中断", completed.stderr)
        self.assertEqual(list(self.output_dir.iterdir()), [])
        self.assertEqual(self.upload_file.read_bytes(), original)

    def test_result_file_write_failure_keeps_input_and_never_reports_success(self) -> None:
        example = self.load_example()
        original = self.upload_file.read_bytes()
        output = self.output_dir / "结果.xlsx"
        original_open = Path.open
        attempted = []

        def reject_output(path, *args, **kwargs):
            if path == output:
                attempted.append(path)
                raise PermissionError("模拟结果目录不可写")
            return original_open(path, *args, **kwargs)

        with patch.object(Path, "open", reject_output):
            completed = self.run_main(example)
        self.assertTrue(attempted, "必须实际走到结果文件写入阶段")
        self.assert_failed(completed)
        self.assert_input_unchanged(self.upload_file, original)
        self.assertEqual({path.name for path in self.output_dir.iterdir()}, {"输入.xlsx"})


if __name__ == "__main__":
    unittest.main()
