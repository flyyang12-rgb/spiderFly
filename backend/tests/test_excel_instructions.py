from __future__ import annotations

import io
import json
import os
import subprocess
import sys
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from unittest.mock import Mock, patch
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook

from spiderfly_instructions import InstructionError, InstructionRegistry
from spiderfly_instructions.excel import INSTRUCTION_ID, READ_EXCEL


class ReadExcelInstructionTests(unittest.TestCase):
    def setUp(self) -> None:
        directory = tempfile.TemporaryDirectory(prefix="spiderfly-excel-tests-")
        self.addCleanup(directory.cleanup)
        self.directory = Path(directory.name)
        self.registry = InstructionRegistry()
        self.registry.register(READ_EXCEL)

    def save_book(self, workbook: Workbook, name: str = "sample.xlsx") -> Path:
        path = self.directory / name
        try:
            workbook.save(path)
        finally:
            workbook.close()
        return path

    def make_table(self, rows: list, name: str = "sample.xlsx") -> Path:
        workbook = Workbook()
        workbook.active.title = "数据"
        for row in rows:
            workbook.active.append(row)
        return self.save_book(workbook, name)

    def read(self, path: Path, **parameters):
        return self.registry.execute(
            INSTRUCTION_ID, {"file_path": str(path), **parameters}
        )

    def assert_read_error(self, path: Path, code: str, **parameters) -> InstructionError:
        with self.assertRaises(InstructionError) as caught:
            self.read(path, **parameters)
        error = caught.exception
        self.assertEqual(error.code, code)
        self.assertEqual(error.instruction_id, INSTRUCTION_ID)
        self.assertEqual(error.stage, "execute")
        self.assertTrue(error.to_dict()["message"])
        return error

    def test_reads_default_and_named_sheets_without_converting_cell_values(self) -> None:
        timestamp = datetime(2026, 9, 3, 14, 5, 6)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "明细"
        sheet.append([" 订单号 ", "数量", "启用", "日期", "备注"])
        sheet.append(["00123", 0, False, timestamp, None])
        sheet.append([123, 2.5, True, None, "  保留空格  "])
        sheet["A3"].number_format = "00000"
        summary = workbook.create_sheet("汇总")
        summary.append(["状态"])
        summary.append(["完成"])
        path = self.save_book(workbook)

        result = self.read(path, required_columns=[" 订单号 ", "日期"])
        self.assertEqual(result.sheet_name, "明细")
        self.assertEqual(result.columns, ["订单号", "数量", "启用", "日期", "备注"])
        self.assertEqual(result.row_count, 2)
        self.assertEqual(result.rows[0], {
            "订单号": "00123", "数量": 0, "启用": False,
            "日期": timestamp, "备注": None,
        })
        self.assertIs(type(result.rows[0]["数量"]), int)
        self.assertIs(result.rows[0]["启用"], False)
        self.assertEqual(result.rows[1]["订单号"], 123)
        self.assertIs(type(result.rows[1]["订单号"]), int)
        self.assertEqual(result.rows[1]["数量"], 2.5)
        self.assertEqual(result.rows[1]["备注"], "  保留空格  ")
        self.assertEqual(self.read(path, sheet_name="汇总").model_dump(), {
            "sheet_name": "汇总", "columns": ["状态"],
            "rows": [{"状态": "完成"}], "row_count": 1,
        })

    def test_skips_empty_rows_and_formatting_only_tails_but_keeps_zero_and_false(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "数据"
        for row in (["值", "备注"], [None, None], ["", ""], [0], [False], [" "]):
            sheet.append(row)
        sheet["H1"].number_format = "0.00"
        sheet["H40"].number_format = "0.00"
        result = self.read(self.save_book(workbook))
        self.assertEqual(result.columns, ["值", "备注"])
        self.assertEqual(result.row_count, 3)
        self.assertEqual(result.rows, [
            {"值": 0, "备注": None},
            {"值": False, "备注": None},
            {"值": " ", "备注": None},
        ])
        self.assertIs(type(result.rows[0]["值"]), int)
        self.assertIs(result.rows[1]["值"], False)

    def test_header_only_sheet_returns_an_empty_table(self) -> None:
        result = self.read(self.make_table([["订单号", "数量"]]))
        self.assertEqual(result.columns, ["订单号", "数量"])
        self.assertEqual(result.rows, [])
        self.assertEqual(result.row_count, 0)

    def test_reports_missing_columns_sheet_and_invalid_required_column(self) -> None:
        path = self.make_table([["订单号"], ["00123"]])
        cases = [
            ({"required_columns": ["客户编号"]}, "EXCEL_COLUMNS_MISSING", "客户编号"),
            ({"sheet_name": "不存在的表"}, "EXCEL_SHEET_MISSING", "不存在的表"),
            ({"required_columns": ["  "]}, "EXCEL_COLUMNS_INVALID", "必需列名"),
        ]
        for parameters, code, expected_detail in cases:
            with self.subTest(parameters=parameters):
                error = self.assert_read_error(path, code, **parameters)
                self.assertIn(expected_detail, error.to_dict()["message"])

    def test_reports_missing_file_unsupported_extension_and_invalid_workbook(self) -> None:
        self.assert_read_error(self.directory / "missing.xlsx", "EXCEL_FILE_MISSING")
        self.assert_read_error(self.directory / "legacy.xls", "EXCEL_FORMAT_UNSUPPORTED")
        invalid = self.directory / "invalid.xlsx"
        invalid.write_bytes(b"this is not an xlsx archive")
        self.assert_read_error(invalid, "EXCEL_FILE_INVALID")

    def test_rejects_empty_duplicate_nontext_headers_and_unheaded_data(self) -> None:
        cases = {
            "empty": [],
            "gap": [["订单号", None, "数量"], ["00123", None, 2]],
            "duplicate": [[" 订单号", "订单号 "], ["00123", "00124"]],
            "nontext": [["订单号", 123], ["00123", 2]],
            "outside": [["订单号"], ["00123", 2]],
        }
        for label, rows in cases.items():
            with self.subTest(case=label):
                self.assert_read_error(
                    self.make_table(rows, f"{label}.xlsx"), "EXCEL_HEADER_INVALID"
                )

    def test_rejects_formulas_and_excel_errors_with_cell_location(self) -> None:
        for label, value in (("formula", "=1+1"), ("error", "#DIV/0!")):
            with self.subTest(case=label):
                path = self.make_table([["订单号", "数量"], ["00123", value]], f"{label}.xlsx")
                error = self.assert_read_error(path, "EXCEL_CELL_UNSUPPORTED")
                self.assertIn("数据", error.to_dict()["message"])
                self.assertIn("B2", error.to_dict()["message"])

    def test_reads_all_data_when_exported_dimensions_are_incorrect(self) -> None:
        path = self.make_table([["订单号", "数量"], ["00123", 2], ["00124", 3]])
        original = path.read_bytes()
        with ZipFile(io.BytesIO(original)) as source, ZipFile(path, "w", ZIP_DEFLATED) as target:
            for item in source.infolist():
                content = source.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    root = ElementTree.fromstring(content)
                    dimension = root.find("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}dimension")
                    self.assertIsNotNone(dimension)
                    dimension.set("ref", "A1:A1")
                    content = ElementTree.tostring(root, encoding="utf-8")
                target.writestr(item, content)
        result = self.read(path)
        self.assertEqual(result.columns, ["订单号", "数量"])
        self.assertEqual(result.rows, [
            {"订单号": "00123", "数量": 2},
            {"订单号": "00124", "数量": 3},
        ])

    def test_success_and_failure_close_resources_without_changing_the_file(self) -> None:
        for label, value, expected_error in (
            ("success", "00123", None),
            ("failure", "=1+1", "EXCEL_CELL_UNSUPPORTED"),
        ):
            with self.subTest(case=label):
                path = self.make_table([["订单号"], [value]], f"{label}.xlsx")
                before = path.read_bytes()
                opened = []

                def observe_load(source, **options):
                    workbook = load_workbook(source, **options)
                    close = Mock(wraps=workbook.close)
                    workbook.close = close
                    opened.append((source, close))
                    return workbook

                with patch("spiderfly_instructions.excel.load_workbook", side_effect=observe_load):
                    if expected_error:
                        self.assert_read_error(path, expected_error)
                    else:
                        self.read(path)
                self.assertEqual(len(opened), 1)
                source, close = opened[0]
                close.assert_called_once()
                self.assertTrue(source.closed)
                self.assertEqual(path.read_bytes(), before)
                renamed = path.with_name(f"{label}-closed.xlsx")
                path.rename(renamed)
                self.assertEqual(renamed.read_bytes(), before)

    def test_cli_prints_json_for_a_success_and_a_known_failure(self) -> None:
        path = self.make_table([["订单号", "日期"], ["00123", datetime(2026, 9, 3)]])
        command = [sys.executable, "-m", "spiderfly_instructions.excel", str(path)]
        options = {
            "cwd": Path(__file__).resolve().parents[1],
            "capture_output": True, "text": True, "encoding": "utf-8",
            "env": {**os.environ, "PYTHONIOENCODING": "utf-8"}, "timeout": 15,
        }
        success = subprocess.run(command + ["--require", "订单号"], **options)
        self.assertEqual(success.returncode, 0, success.stderr)
        self.assertEqual(success.stderr, "")
        data = json.loads(success.stdout)
        self.assertEqual(data["row_count"], 1)
        self.assertEqual(data["rows"], [{"订单号": "00123", "日期": "2026-09-03T00:00:00"}])
        failure = subprocess.run(command + ["--sheet", "不存在"], **options)
        self.assertEqual(failure.returncode, 2)
        self.assertEqual(failure.stdout, "")
        self.assertEqual(json.loads(failure.stderr)["code"], "EXCEL_SHEET_MISSING")


if __name__ == "__main__":
    unittest.main()
