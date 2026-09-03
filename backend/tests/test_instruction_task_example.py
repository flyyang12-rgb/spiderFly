from __future__ import annotations

import importlib.util
import io
import json
import os
import subprocess
import sys
import tempfile
import unittest
from contextlib import redirect_stderr, redirect_stdout
from hashlib import sha256
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook


BACKEND = Path(__file__).resolve().parents[1]
EXAMPLE = BACKEND.parent / "examples" / "instruction_excel_name_task.py"


class InstructionTaskExampleTests(unittest.TestCase):
    def setUp(self) -> None:
        directory = tempfile.TemporaryDirectory(prefix="spiderfly-instruction-task-")
        self.addCleanup(directory.cleanup)
        self.directory = Path(directory.name)
        self.working_dir = self.directory / "working"
        self.working_dir.mkdir()
        self.make_execution("execution")
        self.upload_file = self.make_input("上传名单.xlsx", ["编号", "姓", "名", "启用", "数量", "备注"], [
            ["0007", " 欧阳 ", " 娜娜 ", False, 0, "  原样保留  "],
            ["0420", "司马", None, True, 2.5, "第二条"],
        ])

    def make_execution(self, name: str) -> None:
        self.artifact_dir = self.directory / name / "artifacts"
        self.artifact_dir.mkdir(parents=True)
        self.result_file = self.artifact_dir.parent / "result.json"

    def make_input(self, name: str, columns: list[str], rows: list[list]) -> Path:
        path = self.directory / name
        workbook = Workbook()
        try:
            workbook.active.title = "上传名单"
            workbook.active.append(columns)
            for row in rows:
                workbook.active.append(row)
            workbook.save(path)
        finally:
            workbook.close()
        return path

    def run_example(
        self, *, missing: str | None = None, without_package: bool = False,
        template_path: Path | None = None,
    ):
        environment = {
            key: value for key, value in os.environ.items()
            if not key.startswith("SPIDERFLY_")
        }
        environment.update({
            "PYTHONIOENCODING": "utf-8",
            "PYTHONPATH": "" if without_package else str(BACKEND),
            "SPIDERFLY_RESULT_FILE": str(self.result_file),
            "SPIDERFLY_ARTIFACT_DIR": str(self.artifact_dir),
            "SPIDERFLY_TEMPLATE_FILE": str(template_path or self.upload_file),
        })
        if missing is not None:
            environment.pop(missing)
        arguments = [sys.executable]
        if without_package:
            arguments.append("-S")
        return subprocess.run(
            [*arguments, str(EXAMPLE)], cwd=self.working_dir, env=environment,
            capture_output=True, text=True, encoding="utf-8", timeout=15,
        )

    def receipt(self) -> dict:
        return json.loads(self.result_file.read_text(encoding="utf-8"))

    @staticmethod
    def rows(path: Path) -> list[tuple]:
        workbook = load_workbook(path, read_only=True)
        try:
            return list(workbook["上传名单"].values)
        finally:
            workbook.close()

    def assert_failed(self, completed) -> None:
        self.assertNotEqual(completed.returncode, 0)
        receipt = self.receipt()
        self.assertEqual(receipt["schema_version"], 1)
        self.assertEqual(receipt["outcome"], "failure")
        self.assertEqual(receipt["code"], "EXCEL_NAME_FAILED")
        self.assertIs(receipt["retryable"], False)
        self.assertLessEqual(len(receipt["message"]), 1000)
        self.assertNotIn("已合并", completed.stdout)
        self.assertEqual(list(self.working_dir.iterdir()), [])

    def test_uploaded_example_creates_readable_results_and_platform_receipt(self) -> None:
        original_hash = sha256(self.upload_file.read_bytes()).hexdigest()
        completed = self.run_example()
        self.assertEqual(completed.returncode, 0, completed.stderr)
        self.assertEqual(completed.stderr, "")
        folder = self.artifact_dir / "姓名合并示例"
        output = folder / "结果.xlsx"
        self.assertEqual(self.rows(folder / "输入.xlsx"), [
            ("编号", "姓", "名", "启用", "数量", "备注"),
            ("0007", " 欧阳 ", " 娜娜 ", False, 0, "  原样保留  "),
            ("0420", "司马", None, True, 2.5, "第二条"),
        ])
        output_rows = self.rows(output)
        self.assertEqual(output_rows, [
            ("编号", "姓", "名", "启用", "数量", "备注", "姓名"),
            ("0007", " 欧阳 ", " 娜娜 ", False, 0, "  原样保留  ", "欧阳娜娜"),
            ("0420", "司马", None, True, 2.5, "第二条", "司马"),
        ])
        self.assertIs(output_rows[1][3], False)
        self.assertIs(type(output_rows[1][4]), int)
        self.assertEqual(sha256(self.upload_file.read_bytes()).hexdigest(), original_hash)
        self.assertEqual(sha256((folder / "输入.xlsx").read_bytes()).hexdigest(), original_hash)
        expected_message = f"已合并 2 行姓名，结果文件：{output}"
        self.assertEqual(self.receipt(), {
            "schema_version": 1, "outcome": "success", "code": "EXCEL_NAME_DONE",
            "message": expected_message, "retryable": False,
        })
        self.assertIn(expected_message, completed.stdout)
        self.assertEqual(list(self.working_dir.iterdir()), [])
        self.assertEqual(list(self.result_file.parent.glob(".result-*.tmp")), [])

    def test_missing_upload_never_falls_back_to_demo_data(self) -> None:
        completed = self.run_example(missing="SPIDERFLY_TEMPLATE_FILE")
        self.assert_failed(completed)
        self.assertIn("上传", completed.stderr)
        self.assertEqual(list(self.artifact_dir.iterdir()), [])

    def test_invalid_upload_paths_and_format_never_produce_a_result(self) -> None:
        wrong_format = self.directory / "名单.csv"
        wrong_format.write_bytes(self.upload_file.read_bytes())
        directory_input = self.directory / "文件夹.xlsx"
        directory_input.mkdir()
        corrupt = self.directory / "损坏.xlsx"
        corrupt.write_bytes(b"This is not an Excel workbook")
        cases = [
            ("missing", self.directory / "不存在.xlsx"),
            ("wrong-format", wrong_format),
            ("directory", directory_input),
            ("corrupt", corrupt),
        ]
        for label, source in cases:
            with self.subTest(case=label):
                self.make_execution(f"invalid-{label}")
                original = source.read_bytes() if source.is_file() else None
                completed = self.run_example(template_path=source)
                self.assert_failed(completed)
                self.assertFalse((self.artifact_dir / "姓名合并示例" / "结果.xlsx").exists())
                if label != "corrupt":
                    self.assertEqual(list(self.artifact_dir.iterdir()), [])
                if original is not None:
                    self.assertEqual(source.read_bytes(), original)

    def test_invalid_uploaded_columns_and_name_values_fail_without_output(self) -> None:
        cases = [
            ("missing-surname", ["名"], [["小雨"]], "缺少必需列"),
            ("missing-given-name", ["姓"], [["赵"]], "缺少必需列"),
            ("existing-full-name", ["姓", "名", "姓名"], [["赵", "小雨", "原名"]], "已有“姓名”列"),
            ("number", ["姓", "名"], [["赵", "小雨"], [123, "六"]], "只能填写文字"),
            ("boolean", ["姓", "名"], [["赵", "小雨"], ["钱", False]], "只能填写文字"),
        ]
        for label, columns, rows, expected_message in cases:
            with self.subTest(case=label):
                self.make_execution(f"bad-data-{label}")
                source = self.make_input(f"{label}.xlsx", columns, rows)
                original = source.read_bytes()
                completed = self.run_example(template_path=source)
                self.assert_failed(completed)
                self.assertIn(expected_message, completed.stderr)
                self.assertFalse((self.artifact_dir / "姓名合并示例" / "结果.xlsx").exists())
                self.assertEqual(source.read_bytes(), original)
                self.assertEqual((self.artifact_dir / "姓名合并示例" / "输入.xlsx").read_bytes(), original)

    def test_repeated_execution_fails_without_changing_existing_artifacts(self) -> None:
        first = self.run_example()
        self.assertEqual(first.returncode, 0, first.stderr)
        before = {
            path: path.read_bytes() for path in self.artifact_dir.rglob("*") if path.is_file()
        }
        repeated = self.run_example()
        self.assert_failed(repeated)
        self.assertEqual({
            path: path.read_bytes() for path in self.artifact_dir.rglob("*") if path.is_file()
        }, before)

    def test_interrupted_input_copy_cleans_partial_file_and_never_runs_flow(self) -> None:
        spec = importlib.util.spec_from_file_location("instruction_task_copy_failure", EXAMPLE)
        example = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(example)
        original = self.upload_file.read_bytes()
        stdout, stderr = io.StringIO(), io.StringIO()

        def interrupted_copy(incoming, outgoing) -> None:
            outgoing.write(incoming.read(16))
            outgoing.flush()
            raise OSError("模拟磁盘写入中断")

        with (
            patch.dict(os.environ, {
                "SPIDERFLY_RESULT_FILE": str(self.result_file),
                "SPIDERFLY_ARTIFACT_DIR": str(self.artifact_dir),
                "SPIDERFLY_TEMPLATE_FILE": str(self.upload_file),
            }, clear=True),
            patch.object(example.shutil, "copyfileobj", side_effect=interrupted_copy) as copy,
            patch("example_flows.excel_name.run_flow") as run_flow,
            redirect_stdout(stdout), redirect_stderr(stderr),
        ):
            exit_code = example.main()

        copy.assert_called_once()
        run_flow.assert_not_called()
        completed = subprocess.CompletedProcess([], exit_code, stdout.getvalue(), stderr.getvalue())
        self.assert_failed(completed)
        self.assertIn("模拟磁盘写入中断", completed.stderr)
        self.assertEqual(list((self.artifact_dir / "姓名合并示例").iterdir()), [])
        self.assertEqual(list(self.result_file.parent.glob(".result-*.tmp")), [])
        self.assertEqual(self.upload_file.read_bytes(), original)

    def test_missing_result_variable_does_not_create_files(self) -> None:
        completed = self.run_example(missing="SPIDERFLY_RESULT_FILE")
        self.assertNotEqual(completed.returncode, 0)
        self.assertIn("SPIDERFLY_RESULT_FILE", completed.stderr)
        self.assertFalse(self.result_file.exists())
        self.assertEqual(list(self.artifact_dir.iterdir()), [])
        self.assertEqual(list(self.working_dir.iterdir()), [])

    def test_missing_artifact_variable_writes_failure_only_at_platform_result_path(self) -> None:
        completed = self.run_example(missing="SPIDERFLY_ARTIFACT_DIR")
        self.assert_failed(completed)
        self.assertIn("SPIDERFLY_ARTIFACT_DIR", completed.stderr)
        self.assertEqual(self.receipt()["outcome"], "failure")
        self.assertEqual(list(self.artifact_dir.iterdir()), [])
        self.assertEqual(list(self.working_dir.iterdir()), [])

    def test_missing_instruction_package_has_readable_failure_receipt(self) -> None:
        completed = self.run_example(without_package=True)
        self.assert_failed(completed)
        self.assertIn("spiderfly-instructions", completed.stderr)
        self.assertEqual(self.receipt()["outcome"], "failure")
        self.assertEqual(list(self.artifact_dir.iterdir()), [])
        self.assertEqual(list(self.working_dir.iterdir()), [])


if __name__ == "__main__":
    unittest.main()
