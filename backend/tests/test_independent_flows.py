from __future__ import annotations

import ast
import hashlib
import json
import os
import shutil
import subprocess
import sys
import tempfile
import unittest
from importlib.metadata import version
from pathlib import Path

from openpyxl import Workbook, load_workbook


FLOWS = Path(__file__).resolve().parents[2] / "flows"
# Each child can import only the installed public library, never bundled business examples.
CHILD = """
import importlib.abc, json, runpy, sys
from pathlib import Path
from importlib.metadata import version
class Boundary(importlib.abc.MetaPathFinder):
    def find_spec(self, fullname, path=None, target=None):
        if fullname.partition('.')[0] in {'example_flows', 'app', 'flows', 'examples'}:
            raise ImportError('Independent flow imported a forbidden module: ' + fullname)
sys.meta_path.insert(0, Boundary())
from spiderfly_instructions import InstructionRegistry
actual = InstructionRegistry.execute
calls = []
def trace(registry, name, inputs=None):
    calls.append(name)
    return actual(registry, name, inputs)
InstructionRegistry.execute = trace
record = Path(sys.argv[1])
sys.argv = sys.argv[2:]
try:
    runpy.run_path(sys.argv[0], run_name='__main__')
finally:
    record.write_text(json.dumps({
        'calls': calls, 'version': version('spiderfly-instructions'),
        'forbidden_loaded': [n for n in sys.modules if n.partition('.')[0] in {'example_flows','app','flows','examples'}],
    }), encoding='utf-8')
"""


class IndependentFlowTests(unittest.TestCase):
    def setUp(self):
        directory = tempfile.TemporaryDirectory(prefix="spiderfly-flow-boundary-")
        self.addCleanup(directory.cleanup)
        self.root = Path(directory.name)
        self.sequence = 0
        self.source = self.root / "input.xlsx"
        book = Workbook()
        try:
            sheet = book.active
            sheet.title = "订单"
            for row in [
                ["订单号", "状态", "金额"],
                ["001", "待处理", "5,2"],
                ["002", "已完成", 12.3],
                ["003", "待处理", 5.6],
            ]:
                sheet.append(row)
            book.create_sheet("说明")["A1"] = "保留其他页"
            book.save(self.source)
        finally:
            book.close()
        self.original = self.source.read_bytes()

    def invoke(self, filename, args=(), *, platform=False, upload=True, changed_source=None):
        self.sequence += 1
        directory = self.root / f"run_{self.sequence}"
        entry_dir = directory / "entry"
        entry_dir.mkdir(parents=True)
        entry = entry_dir / filename
        if changed_source is None:
            shutil.copy2(FLOWS / filename, entry)
        else:
            entry.write_text(changed_source, encoding="utf-8")
        self.assertEqual(list(entry_dir.iterdir()), [entry])
        environment = {key: value for key, value in os.environ.items()
                       if not key.upper().startswith(("SPIDERFLY_", "FEISHU_"))
                       and key.upper() not in ("PYTHONPATH", "PYTHONHOME")}
        if platform:
            (directory / "artifacts").mkdir()
            environment.update({
                "SPIDERFLY_RESULT_FILE": str(directory / "result.json"),
                "SPIDERFLY_ARTIFACT_DIR": str(directory / "artifacts"),
            })
            if upload:
                environment["SPIDERFLY_TEMPLATE_FILE"] = str(self.source)
        result = subprocess.run(
            [sys.executable, "-I", "-X", "utf8", "-c", CHILD, str(directory / "trace.json"), str(entry), *map(str, args)],
            cwd=entry_dir, env=environment, capture_output=True, text=True, encoding="utf-8", timeout=20,
        )
        trace = json.loads((directory / "trace.json").read_text(encoding="utf-8"))
        self.assertEqual(trace["version"], version("spiderfly-instructions"))
        self.assertEqual(trace["forbidden_loaded"], [])
        self.assertEqual(self.source.read_bytes(), self.original)
        return result, directory, trace["calls"]

    def check_average(self, path):
        book = load_workbook(path)
        try:
            self.assertEqual(book.sheetnames, ["订单", "说明"])
            self.assertEqual(list(book["订单"].values), [
                ("订单号", "状态", "金额", "待处理平均数"),
                ("001", "待处理", "5,2", 3.5),
                ("002", "已完成", 12.3, None),
                ("003", "待处理", 5.6, 5.6),
            ])
            self.assertEqual(book["说明"]["A1"].value, "保留其他页")
            self.assertEqual(book["订单"]["A2"].data_type, "s")
        finally:
            book.close()

    def test_amount_cli_calls_public_instructions_with_business_imports_blocked(self):
        result, _, calls = self.invoke("amount_difference.py", [self.source])
        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertEqual(json.loads(result.stdout)["difference"], "0.3")
        self.assertEqual(calls, ["excel.read", "table.filter_equals", "table.filter_equals"])

    def test_average_cli_calls_public_instructions_and_preserves_input(self):
        output = self.root / "average.xlsx"
        result, _, calls = self.invoke("pending_average.py", [self.source, output])
        self.assertEqual(result.returncode, 0, result.stderr)
        self.check_average(output)
        self.assertEqual(calls, ["excel.read", "math.average", "math.average", "excel.write"])

    def test_single_uploaded_file_runs_both_platform_flows(self):
        for name, code in (("amount_difference.py", "AMOUNT_DIFFERENCE_DONE"),
                           ("pending_average.py", "EXCEL_AVERAGE_DONE")):
            with self.subTest(name=name):
                result, directory, _ = self.invoke(name, platform=True)
                self.assertEqual(result.returncode, 0, result.stderr)
                receipt = json.loads((directory / "result.json").read_text(encoding="utf-8"))
                self.assertEqual((receipt["outcome"], receipt["code"]), ("success", code))
                self.assertEqual((directory / "artifacts/流程文件/输入.xlsx").read_bytes(), self.original)
                if name == "pending_average.py":
                    self.check_average(directory / "artifacts/流程文件/输出/结果.xlsx")
                else:
                    self.assertIn("差额：0.3", (directory / "artifacts/流程文件/输出/金额差额.txt").read_text(encoding="utf-8"))

    def test_changing_only_flow_changes_business_without_library_changes(self):
        import spiderfly_instructions
        library = Path(spiderfly_instructions.__file__).parent
        before = {path.name: hashlib.sha256(path.read_bytes()).hexdigest() for path in library.glob("*.py")}
        source = (FLOWS / "amount_difference.py").read_text(encoding="utf-8")
        old = 'LEFT_STATUS = "待处理"\nRIGHT_STATUS = "已完成"'
        self.assertIn(old, source)
        changed = source.replace(old, 'LEFT_STATUS = "已完成"\nRIGHT_STATUS = "待处理"')
        result, _, _ = self.invoke("amount_difference.py", [self.source], changed_source=changed)
        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertEqual(json.loads(result.stdout)["difference"], "-0.3")
        result, directory, _ = self.invoke("amount_difference.py", platform=True, changed_source=changed)
        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertIn("差额：-0.3", (directory / "artifacts/流程文件/输出/金额差额.txt").read_text(encoding="utf-8"))
        result, _, _ = self.invoke("amount_difference.py", [self.source])
        self.assertEqual(json.loads(result.stdout)["difference"], "0.3")
        self.assertEqual(before, {path.name: hashlib.sha256(path.read_bytes()).hexdigest() for path in library.glob("*.py")})

    def test_invalid_data_does_not_succeed_or_write_results(self):
        book = load_workbook(self.source)
        try:
            book["订单"]["C2"] = "bad"
            book.save(self.source)
        finally:
            book.close()
        self.original = self.source.read_bytes()
        for name, code in (("amount_difference.py", "FLOW_AMOUNT_INVALID"),
                           ("pending_average.py", "MATH_VALUE_INVALID")):
            with self.subTest(name=name):
                result, directory, _ = self.invoke(name, platform=True)
                self.assertEqual(result.returncode, 1)
                receipt = json.loads((directory / "result.json").read_text(encoding="utf-8"))
                self.assertEqual((receipt["outcome"], receipt["code"]), ("failure", code))
                self.assertEqual(list((directory / "artifacts/流程文件/输出").iterdir()), [])

    def test_template_is_independent_and_has_no_business_package_imports(self):
        for path in FLOWS.glob("*.py"):
            tree = ast.parse(path.read_text(encoding="utf-8"))
            for node in ast.walk(tree):
                if isinstance(node, ast.ImportFrom):
                    self.assertEqual(node.level, 0)
                    self.assertNotIn((node.module or "").split(".")[0], {"example_flows", "app", "flows", "examples"})
        result, directory, calls = self.invoke("template.py", platform=True, upload=False)
        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertEqual(calls, [])
        self.assertEqual(json.loads((directory / "result.json").read_text(encoding="utf-8"))["outcome"], "success")


if __name__ == "__main__":
    unittest.main()
