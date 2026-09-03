from __future__ import annotations

import tempfile
import unittest
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from spiderfly_instructions import InstructionError, InstructionRegistry
from spiderfly_instructions.excel import READ_EXCEL
from spiderfly_instructions.excel_write import (
    INSTRUCTION_ID,
    WRITE_EXCEL,
    WriteExcelInput,
    verify_excel_write,
)


class PartialWriteFailure:
    """Create real partial output, then simulate a failed destination write."""

    def __init__(self, stream, error: BaseException) -> None:
        self.stream = stream
        self.error = error

    def __enter__(self):
        return self

    def __exit__(self, *exception_details):
        self.stream.close()
        return False

    def write(self, data: bytes) -> int:
        self.stream.write(data[:16])
        self.stream.flush()
        raise self.error


class WriteExcelInstructionTests(unittest.TestCase):
    def setUp(self) -> None:
        directory = tempfile.TemporaryDirectory(prefix="spiderfly-excel-write-tests-")
        self.addCleanup(directory.cleanup)
        self.directory = Path(directory.name)
        self.path = self.directory / "output.xlsx"
        self.registry = InstructionRegistry()
        self.registry.register(WRITE_EXCEL)
        self.registry.register(READ_EXCEL)

    def parameters(self, **changes) -> dict:
        values = {
            "file_path": str(self.path),
            "columns": ["订单号"],
            "rows": [{"订单号": "00123"}],
        }
        values.update(changes)
        return values

    def write(self, **changes):
        return self.registry.execute(INSTRUCTION_ID, self.parameters(**changes))

    def assert_write_error(self, code: str, **changes) -> InstructionError:
        with self.assertRaises(InstructionError) as caught:
            self.write(**changes)
        error = caught.exception
        self.assertEqual(error.code, code)
        self.assertEqual(error.instruction_id, INSTRUCTION_ID)
        self.assertEqual(error.stage, "input" if code == "INPUT_INVALID" else "execute")
        self.assertTrue(error.to_dict()["message"])
        return error

    def test_round_trip_preserves_column_order_literal_strings_and_supported_values(self) -> None:
        columns = ["订单号", "=原文", "错误文字", "数量", "启用", "金额", "日期", "时间点", "时刻", "时长", "空值"]
        # Dict insertion order intentionally differs from the specified columns.
        row = {
            "空值": "", "时长": timedelta(days=1, seconds=2, milliseconds=123),
            "时刻": time(14, 5, 6, 123000), "时间点": datetime(2026, 9, 3, 14, 5, 6, 123000),
            "日期": date(2026, 9, 3), "金额": 2.5, "启用": False, "数量": 0,
            "错误文字": "#N/A", "=原文": "=1+1", "订单号": "00123",
        }
        result = self.write(columns=columns, rows=[row], sheet_name="明细")
        self.assertEqual(result.file_path, str(self.path.absolute()))
        self.assertEqual(result.sheet_name, "明细")
        self.assertEqual(result.row_count, 1)
        returned = self.registry.execute("excel.read", {"file_path": result.file_path})
        self.assertEqual(returned.columns, columns)
        self.assertEqual(returned.rows, [{**row, "空值": None}])
        self.assertIs(type(returned.rows[0]["数量"]), int)
        self.assertIs(returned.rows[0]["启用"], False)
        self.assertIs(type(returned.rows[0]["日期"]), date)
        with self.path.open("rb") as source:
            workbook = load_workbook(source, read_only=True, data_only=False)
            try:
                sheet = workbook["明细"]
                self.assertEqual([cell.value for cell in sheet[1]], columns)
                for address in ("B1", "A2", "B2", "C2"):
                    self.assertEqual(sheet[address].data_type, "s")
            finally:
                workbook.close()

    def test_header_only_table_and_fully_empty_rows_have_explicit_counts(self) -> None:
        for label, rows in (
            ("header-only", []),
            ("blank-rows", [{"甲": None, "乙": None}, {"甲": "", "乙": ""}]),
        ):
            with self.subTest(case=label):
                path = self.directory / f"{label}.xlsx"
                result = self.write(file_path=str(path), columns=["甲", "乙"], rows=rows)
                self.assertEqual(result.sheet_name, "数据")
                self.assertEqual(result.row_count, len(rows))
                with path.open("rb") as source:
                    workbook = load_workbook(source, read_only=True)
                    try:
                        sheet = workbook["数据"]
                        self.assertEqual(sheet.max_row, len(rows) + 1)
                        self.assertEqual(sheet.max_column, 2)
                        self.assertEqual(list(sheet.values), [("甲", "乙")] + [(None, None)] * len(rows))
                    finally:
                        workbook.close()
                read_result = self.registry.execute("excel.read", {"file_path": str(path)})
                self.assertEqual(read_result.rows, [])
                self.assertEqual(read_result.row_count, 0)

    def test_existing_output_is_never_overwritten_or_deleted(self) -> None:
        self.write()
        original = self.path.read_bytes()
        self.assert_write_error("EXCEL_FILE_EXISTS", rows=[{"订单号": "different"}])
        self.assertEqual(self.path.read_bytes(), original)

    def test_bad_headers_rows_and_sheet_names_do_not_create_output(self) -> None:
        cases = [
            ("INPUT_INVALID", {"columns": [], "rows": []}),
            ("EXCEL_HEADER_INVALID", {"columns": ["订单号", "订单号"]}),
            ("EXCEL_HEADER_INVALID", {"columns": [""]}),
            ("EXCEL_HEADER_INVALID", {"columns": [" 订单号"]}),
            ("EXCEL_ROW_INVALID", {"rows": [{}]}),
            ("EXCEL_ROW_INVALID", {"rows": [{"订单号": "00123", "多余列": 1}]}),
            ("EXCEL_SHEET_INVALID", {"sheet_name": "错误/名称"}),
            ("EXCEL_SHEET_INVALID", {"sheet_name": " "}),
            ("EXCEL_SHEET_INVALID", {"sheet_name": "'名称"}),
            ("INPUT_INVALID", {"sheet_name": "表" * 32}),
        ]
        for code, parameters in cases:
            with self.subTest(parameters=parameters):
                self.assert_write_error(code, **parameters)
                self.assertFalse(self.path.exists())

    def test_values_that_cannot_be_preserved_are_rejected_before_file_creation(self) -> None:
        cases = {
            "nan": float("nan"),
            "infinity": float("inf"),
            "negative-infinity": float("-inf"),
            "large-integer": 10**15,
            "negative-large-integer": -(10**15),
            "long-text": "字" * 32768,
            "control-character": "不能保存\x00",
            "timezone": datetime(2026, 9, 3, tzinfo=timezone.utc),
            "precise-datetime": datetime(2026, 9, 3, microsecond=1),
            "precise-time": time(12, microsecond=1),
            "precise-duration": timedelta(microseconds=1),
        }
        for label, value in cases.items():
            with self.subTest(case=label):
                error = self.assert_write_error("EXCEL_VALUE_INVALID", rows=[{"订单号": value}])
                self.assertIn("第 2 行", error.to_dict()["message"])
                self.assertFalse(self.path.exists())

    def test_missing_directory_and_wrong_extension_leave_no_files(self) -> None:
        missing = self.directory / "not-created" / "output.xlsx"
        self.assert_write_error("EXCEL_DIRECTORY_MISSING", file_path=str(missing))
        self.assertFalse(missing.parent.exists())
        unsupported = self.directory / "output.xls"
        self.assert_write_error("EXCEL_FORMAT_UNSUPPORTED", file_path=str(unsupported))
        self.assertFalse(unsupported.exists())

    def test_partial_write_failure_or_interrupt_closes_and_removes_only_new_output(self) -> None:
        real_open = Path.open
        neighbor = self.directory / "existing.xlsx"
        neighbor.write_bytes(b"existing file must remain unchanged")
        for label, error in (("io-error", OSError("simulated disk failure")), ("interrupt", KeyboardInterrupt())):
            with self.subTest(case=label):
                streams = []

                def fail_after_creating(path, mode="r", *args, **kwargs):
                    stream = real_open(path, mode, *args, **kwargs)
                    if path == self.path and mode == "xb":
                        streams.append(stream)
                        return PartialWriteFailure(stream, error)
                    return stream

                with patch.object(Path, "open", new=fail_after_creating):
                    if isinstance(error, Exception):
                        caught = self.assert_write_error("EXCEL_WRITE_FAILED")
                        self.assertIn("已清理", caught.to_dict()["message"])
                    else:
                        with self.assertRaises(KeyboardInterrupt):
                            self.write()
                self.assertEqual(len(streams), 1)
                self.assertTrue(streams[0].closed)
                self.assertFalse(self.path.exists())
                self.assertEqual(neighbor.read_bytes(), b"existing file must remain unchanged")

    def test_cleanup_failure_reports_the_remaining_partial_output(self) -> None:
        real_open = Path.open
        real_unlink = Path.unlink
        streams = []

        def fail_after_creating(path, mode="r", *args, **kwargs):
            stream = real_open(path, mode, *args, **kwargs)
            if path == self.path and mode == "xb":
                streams.append(stream)
                return PartialWriteFailure(stream, OSError("simulated disk failure"))
            return stream

        def fail_cleanup(path, *args, **kwargs):
            if path == self.path:
                raise PermissionError("simulated cleanup failure")
            return real_unlink(path, *args, **kwargs)

        with patch.object(Path, "open", new=fail_after_creating), patch.object(Path, "unlink", new=fail_cleanup):
            error = self.assert_write_error("EXCEL_WRITE_INCOMPLETE")
        self.assertIn("残留", error.to_dict()["message"])
        self.assertTrue(self.path.exists())
        self.assertEqual(len(self.path.read_bytes()), 16)
        self.assertEqual(len(streams), 1)
        self.assertTrue(streams[0].closed)

    def test_verifier_rejects_incorrect_metadata_and_changed_workbook_structure(self) -> None:
        parameters = self.parameters()
        inputs = WriteExcelInput.model_validate(parameters)
        result = self.registry.execute(INSTRUCTION_ID, parameters)
        self.assertTrue(verify_excel_write(inputs, result))
        for changes in (
            {"file_path": str(self.directory / "other.xlsx")},
            {"sheet_name": "错误名称"},
            {"row_count": 99},
        ):
            with self.subTest(changes=changes):
                self.assertFalse(verify_excel_write(inputs, result.model_copy(update=changes)))
        workbook = load_workbook(self.path)
        try:
            workbook.active["A1"] = "错误表头"
            workbook.save(self.path)
        finally:
            workbook.close()
        self.assertFalse(verify_excel_write(inputs, result))


if __name__ == "__main__":
    unittest.main()
